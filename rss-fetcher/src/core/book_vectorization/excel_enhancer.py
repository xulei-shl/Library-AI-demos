#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel增强器，负责处理Excel文件的读写和列管理

该模块实现了Excel文件的读取、增强和保存功能，主要用于将主题筛选结果
添加到原始的书籍数据Excel文件中，提供完整的数据处理流程。
"""

import os
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from src.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelEnhancer:
    """Excel增强器，负责处理Excel文件的读写和列管理"""
    
    # 主题筛选新增列的命名规范
    THEME_COLUMNS = {
        "is_selected": "初评结果",
        "score": "初评分数",
        "relevance_check": "初评相关性",
        "dimension_match": "初评契合度",
        "reason": "初评理由",
        "llm_status": "初评状态",
        "error_message": "初评错误信息"
    }
    
    def __init__(self, excel_path: str):
        """
        初始化Excel增强器
        
        Args:
            excel_path: Excel文件路径
            
        Raises:
            FileNotFoundError: 文件不存在
            ExcelFormatError: 文件格式错误
        """
        self.excel_path = Path(excel_path)
        
        # 验证文件存在性
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
        
        # 验证文件格式
        if self.excel_path.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"不支持的文件格式: {self.excel_path.suffix}")
        
        logger.info(f"Excel增强器初始化完成: {excel_path}")
        
        # 存储失败的书籍记录
        self._failed_books = []
    
    def load_books_data(self) -> List[Dict]:
        """
        从Excel加载书籍数据
        
        Returns:
            书籍数据列表，每个字典包含完整的书籍元数据
            
        Raises:
            MissingColumnError: 缺少必需的列
        """
        logger.info(f"开始加载Excel数据: {self.excel_path}")
        
        try:
            # 读取Excel文件
            df = pd.read_excel(self.excel_path)
            
            # 检查必需列是否存在
            required_columns = ["书目条码", "豆瓣书名"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                raise ValueError(f"Excel文件缺少必需的列: {missing_columns}")
            
            # 转换为字典列表
            books_data = df.to_dict('records')
            
            logger.info(f"成功加载{len(books_data)}本书籍数据")
            
            return books_data
            
        except Exception as e:
            logger.error(f"加载Excel数据时发生错误: {e}")
            raise
    
    def add_evaluation_results(self, results: List[Dict]) -> str:
        """
        将评估结果添加到Excel文件中
        
        Args:
            results: 评估结果列表，顺序需与书籍数据一致
            
        Returns:
            增强后的Excel文件路径
            
        Raises:
            DataMismatchError: 结果数量与书籍数量不匹配
        """
        logger.info(f"开始添加评估结果到Excel文件")
        
        # 重新加载原始数据
        books_data = self.load_books_data()
        
        # 检查数据匹配
        if len(results) != len(books_data):
            raise ValueError(f"评估结果数量({len(results)})与书籍数量({len(books_data)})不匹配")
        
        try:
            # 创建DataFrame
            df = pd.DataFrame(books_data)
            
            # 添加评估结果列
            df = self._add_evaluation_columns(df, results)
            
            # 生成输出文件名
            output_path = self._generate_output_path()
            
            # 保存增强后的Excel文件
            self._save_enhanced_excel(df, output_path)
            
            logger.info(f"评估结果已添加到Excel文件: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"添加评估结果时发生错误: {e}")
            raise
    
    def get_failed_books(self) -> List[Dict]:
        """
        获取处理失败的书籍列表，用于重试
        
        Returns:
            失败书籍的索引和错误信息列表
        """
        return self._failed_books.copy()
    
    def _add_evaluation_columns(self, df: pd.DataFrame, results: List[Dict]) -> pd.DataFrame:
        """
        将评估结果添加到DataFrame中
        
        Args:
            df: 原始书籍数据DataFrame
            results: 评估结果列表
            
        Returns:
            添加了评估结果的DataFrame
        """
        # 初始化新增列
        for field, column_name in self.THEME_COLUMNS.items():
            df[column_name] = None
        
        # 填充评估结果
        failed_books = []
        
        for i, result in enumerate(results):
            # 获取书籍条码用于匹配
            book_barcode = result.get("book_barcode", "")
            
            # 如果有错误，记录到失败列表
            if result.get("llm_status") == "failed":
                failed_books.append({
                    "index": i,
                    "book_barcode": book_barcode,
                    "error_message": result.get("error_message", "未知错误")
                })
            
            # 填充评估结果
            for field, column_name in self.THEME_COLUMNS.items():
                if field in result:
                    df.at[i, column_name] = result[field]
        
        # 更新失败书籍列表
        self._failed_books = failed_books
        
        return df
    
    def _generate_output_path(self) -> Path:
        """
        生成输出文件路径
        
        Returns:
            输出文件路径
        """
        # 获取原始文件名（不含扩展名）
        stem = self.excel_path.stem
        
        # 添加后缀
        output_stem = f"{stem}_主题筛选增强"
        
        # 构建输出路径
        output_path = self.excel_path.parent / f"{output_stem}.xlsx"
        
        return output_path
    
    def _save_enhanced_excel(self, df: pd.DataFrame, output_path: Path) -> None:
        """
        保存增强后的Excel文件
        
        Args:
            df: 包含评估结果的DataFrame
            output_path: 输出文件路径
        """
        # 确保输出目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 使用openpyxl保存，以便更好地控制格式
        with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='书籍信息', index=False)
            
            # 获取工作表对象，设置列宽
            worksheet = writer.sheets['书籍信息']
            self._adjust_column_width(worksheet, df)
            
            # 设置筛选器
            worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=worksheet.max_column).coordinate}"
        
        logger.info(f"增强后的Excel文件已保存: {output_path}")
    
    def _adjust_column_width(self, worksheet, df):
        """
        调整Excel列宽以适应内容
        
        Args:
            worksheet: Excel工作表对象
            df: DataFrame对象
        """
        from openpyxl.utils import get_column_letter
        
        # 设置最小和最大列宽
        MIN_WIDTH = 10
        MAX_WIDTH = 80  # 主题筛选列可能包含长文本，增加最大宽度
        
        for col_num, column in enumerate(df.columns, 1):
            # 计算列的最大长度
            max_length = max(
                len(str(column)),  # 列名长度
                df[column].astype(str).map(len).max()  # 列中最大值长度
            )
            
            # 限制在最小和最大宽度之间
            adjusted_width = min(max(MIN_WIDTH, max_length), MAX_WIDTH)
            
            # 设置列宽
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 为主题筛选相关列设置特殊格式
        for col_num, column in enumerate(df.columns, 1):
            if "主题筛选" in column:
                column_letter = get_column_letter(col_num)
                
                # 设置文本换行
                for row in range(2, worksheet.max_row + 1):  # 从第二行开始（跳过标题）
                    cell = worksheet[f"{column_letter}{row}"]
                    # 检查单元格是否已有对齐设置
                    if cell.alignment:
                        cell.alignment = cell.alignment.copy(wrap_text=True)
                    else:
                        from openpyxl.styles import Alignment
                        cell.alignment = Alignment(wrap_text=True)
                
                # 设置行高自动调整
                worksheet.row_dimensions.height = 30
    
    def create_summary_report(self, output_path: Optional[str] = None) -> str:
        """
        创建筛选结果摘要报告
        
        Args:
            output_path: 摘要报告输出路径，如果为None则使用默认路径
            
        Returns:
            摘要报告文件路径，如果没有失败记录则返回空字符串
        """
        if not self._failed_books:
            logger.info("没有失败的书籍记录，跳过摘要报告生成")
            return ""
        
        # 生成默认输出路径
        if output_path is None:
            stem = self.excel_path.stem
            output_path = str(self.excel_path.parent / f"{stem}_主题筛选摘要报告.txt")
        
        try:
            # 创建摘要报告
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("主题筛选结果摘要报告\n")
                f.write("=" * 50 + "\n\n")
                
                f.write(f"原始Excel文件: {self.excel_path}\n")
                f.write(f"失败书籍数量: {len(self._failed_books)}\n\n")
                
                f.write("失败书籍详情:\n")
                f.write("-" * 30 + "\n")
                
                for i, failed_book in enumerate(self._failed_books, 1):
                    f.write(f"{i}. 书籍条码: {failed_book['book_barcode']}\n")
                    f.write(f"   错误信息: {failed_book['error_message']}\n\n")
            
            logger.info(f"摘要报告已生成: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"生成摘要报告时发生错误: {e}")
            raise