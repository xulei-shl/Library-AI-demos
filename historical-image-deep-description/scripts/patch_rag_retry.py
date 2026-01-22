#!/usr/bin/env python3
"""
RAG知识库检索重新执行补丁脚本

功能：
1. 清理指定JSON文件中所有实体的RAG检索结果
2. 删除 l3_rag_* 字段和 metadata 中的记录
3. 为后续重新执行RAG检索做准备

使用方式：
    # 清理单个文件
    python scripts/patch_rag_retry.py --row-id 438_002

    # 清理多个文件
    python scripts/patch_rag_retry.py --row-ids 438_002,1861_001,2201_003

    # 清理所有文件（基于Excel）
    python scripts/patch_rag_retry.py --all

    # 清理并指定要清理的任务
    python scripts/patch_rag_retry.py --row-id 438_002 --tasks enhanced_rag_retrieval,web_search
"""

import argparse
import json
import os
import sys
from typing import List, Set, Dict, Any
from pathlib import Path

# 添加项目根目录到路径
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.utils.logger import init_logging, get_logger

try:
    import yaml
except ImportError:
    yaml = None

# 需要清理的RAG相关字段
RAG_FIELDS = {
    "entity_label_retrieval",      # 标准RAG检索
    "enhanced_rag_retrieval",      # 增强RAG检索
    "enhanced_web_retrieval",      # 增强Web检索
    "web_search",                   # Web搜索
}


def load_settings(settings_path: str = "config/settings.yaml") -> Dict[str, Any]:
    """加载配置文件"""
    if yaml is None:
        raise RuntimeError("未安装 pyyaml")
    path = PROJECT_ROOT / settings_path
    if not path.exists():
        raise RuntimeError(f"配置文件不存在：{path}")
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def get_output_dir(settings: Dict[str, Any]) -> str:
    """获取输出目录"""
    # 优先从data.outputs获取，否则从deep_analysis获取
    outputs = settings.get("data", {}).get("outputs", {})
    if outputs.get("enabled", False):
        return outputs.get("dir", "runtime/outputs")
    return settings.get("deep_analysis", {}).get("output_dir", "runtime/outputs")


def get_row_ids_from_excel(excel_path: str, id_column: str = "编号") -> List[str]:
    """从Excel读取所有row_id"""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("未安装 openpyxl，请先安装：pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    # 查找编号列
    headers = [cell.value for cell in ws[1]]
    try:
        id_col_idx = headers.index(id_column)
    except ValueError:
        raise ValueError(f"Excel中没有找到列：{id_column}")

    row_ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_id = str(row[id_col_idx]).strip() if row[id_col_idx] else None
        if row_id:
            row_ids.append(row_id)

    wb.close()
    return row_ids


def clean_entity_rag_fields(entity: Dict[str, Any], tasks: Set[str]) -> bool:
    """
    清理单个实体的RAG字段

    Args:
        entity: 实体节点数据（会被直接修改）
        tasks: 要清理的任务集合

    Returns:
        bool: 是否有字段被清理
    """
    modified = False

    # 清理实体节点中的 l3_rag_* 字段
    for task in tasks:
        field_name = f"l3_rag_{task}"
        if field_name in entity:
            del entity[field_name]
            modified = True

    # 清理 metadata 中的 l3_rag_* 记录
    if "metadata" in entity and isinstance(entity["metadata"], dict):
        metadata = entity["metadata"]
        for task in tasks:
            field_name = f"l3_rag_{task}"
            if field_name in metadata:
                del metadata[field_name]
                modified = True

        # 如果metadata为空，删除它
        if not metadata:
            del entity["metadata"]

    return modified


def clean_json_file(json_path: str, tasks: Set[str], logger) -> bool:
    """
    清理JSON文件中所有实体的RAG字段

    Args:
        json_path: JSON文件路径
        tasks: 要清理的任务集合
        logger: 日志记录器

    Returns:
        bool: 是否有修改
    """
    if not os.path.exists(json_path):
        logger.warning(f"文件不存在，跳过：{json_path}")
        return False

    logger.info(f"开始清理文件：{json_path}")

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    row_id = data.get("row_id", "unknown")
    entities = data.get("entities", [])

    modified_count = 0
    for entity in entities:
        if clean_entity_rag_fields(entity, tasks):
            modified_count += 1

    if modified_count > 0:
        # 写回文件
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"[OK] 已清理 {modified_count}/{len(entities)} 个实体：{json_path}")
        return True
    else:
        logger.info(f"[SKIP] 没有需要清理的字段：{json_path}")
        return False


def main():
    """主函数"""
    init_logging()
    logger = get_logger(__name__)

    parser = argparse.ArgumentParser(
        description="RAG知识库检索重新执行补丁脚本",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--row-id", type=str, help="单个编号（如 438_002）")
    parser.add_argument("--row-ids", type=str, help="多个编号，逗号分隔（如 438_002,1861_001）")
    parser.add_argument("--all", action="store_true", help="处理Excel中的所有编号")
    parser.add_argument("--tasks", type=str, default="all",
                       help="要清理的任务，逗号分隔。默认all（所有RAG任务）")
    parser.add_argument("--settings", type=str, default="config/settings.yaml",
                       help="配置文件路径")
    parser.add_argument("--dry-run", action="store_true",
                       help="仅显示将要清理的内容，不实际修改文件")

    args = parser.parse_args()

    # 解析要清理的任务
    if args.tasks.lower() == "all":
        tasks_to_clean = RAG_FIELDS
    else:
        tasks_to_clean = {t.strip() for t in args.tasks.split(",")}
        # 验证任务名
        invalid = tasks_to_clean - RAG_FIELDS
        if invalid:
            logger.error(f"无效的任务名：{invalid}")
            logger.info(f"有效的任务名：{RAG_FIELDS}")
            return 1

    logger.info(f"准备清理RAG任务：{tasks_to_clean}")

    # 加载配置
    settings = load_settings(args.settings)
    output_dir = get_output_dir(settings)
    logger.info(f"输出目录：{output_dir}")

    # 获取要处理的row_id列表
    row_ids = []
    if args.row_id:
        row_ids = [args.row_id]
    elif args.row_ids:
        row_ids = [rid.strip() for rid in args.row_ids.split(",")]
    elif args.all:
        # 从Excel读取所有编号
        data_config = settings.get("data", {})
        excel_path = data_config.get("paths", {}).get("metadata_excel", "metadata.xlsx")
        id_column = data_config.get("excel", {}).get("columns", {}).get("id", "编号")
        row_ids = get_row_ids_from_excel(excel_path, id_column)
        logger.info(f"从Excel读取到 {len(row_ids)} 个编号")
    else:
        parser.print_help()
        return 1

    # 处理每个JSON文件
    total_files = len(row_ids)
    modified_files = 0

    for row_id in row_ids:
        json_path = os.path.join(output_dir, f"{row_id}.json")

        if args.dry_run:
            # 仅显示将要清理的内容
            logger.info(f"[DRY RUN] 将清理文件：{json_path}")
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                entities = data.get("entities", [])
                for entity in entities:
                    label = entity.get("label", "unknown")
                    for task in tasks_to_clean:
                        field_name = f"l3_rag_{task}"
                        if field_name in entity or (entity.get("metadata", {}).get(field_name)):
                            logger.info(f"  - 实体 '{label}' 的字段 '{field_name}' 将被清理")
        else:
            # 实际清理
            if clean_json_file(json_path, tasks_to_clean, logger):
                modified_files += 1

    # 总结
    if args.dry_run:
        logger.info(f"[DRY RUN] 将处理 {total_files} 个文件")
    else:
        logger.info(f"处理完成：{modified_files}/{total_files} 个文件被修改")

        if modified_files > 0:
            logger.info("\n下一步操作：")
            logger.info("1. 重新执行RAG检索：")
            logger.info("   python main.py --tasks rag+")
            logger.info("\n2. 重新生成Deep分析：")
            logger.info("   python main.py --tasks l3:deep_all --row-id <编号>")

    return 0


if __name__ == "__main__":
    sys.exit(main())
