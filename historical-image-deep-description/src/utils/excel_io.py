import os
import shutil
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .logger import get_logger

logger = get_logger(__name__)

@dataclass
class ExcelConfig:
    sheet_name: str
    columns: Dict[str, str]
    skip_if_present: bool
    create_backup: bool

class ExcelIO:
    def __init__(self, excel_path: str, config: ExcelConfig) -> None:
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel 文件不存在: {excel_path}，请先创建后再运行。")
        self.excel_path = excel_path
        self.config = config
        if self.config.create_backup:
            self._backup()
        self.wb = load_workbook(excel_path)
        self.ws = self._select_sheet(self.wb, self.config.sheet_name)
        self.header_map = self._build_header_map(self.ws)

    def _backup(self) -> None:
        base = os.path.basename(self.excel_path)
        dir_ = os.path.dirname(self.excel_path)
        ts = __import__("time").strftime("%Y%m%d-%H%M%S")
        dst = os.path.join(dir_, f"{base}.bak-{ts}")
        try:
            shutil.copy2(self.excel_path, dst)
            logger.info(f"excel_backup path={dst}")
        except Exception as e:
            logger.warning(f"excel_backup_failed path={dst} err={e}")

    def _select_sheet(self, wb, sheet_name: str) -> Worksheet:
        if sheet_name:
            if sheet_name in wb.sheetnames:
                return wb[sheet_name]
            else:
                raise ValueError(f"未找到工作表: {sheet_name}")
        return wb[wb.sheetnames[0]]

    def _build_header_map(self, ws: Worksheet) -> Dict[str, int]:
        header_map: Dict[str, int] = {}
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            raise ValueError("Excel 首行表头为空。")
        for idx, name in enumerate(first_row, start=1):
            if name is None:
                continue
            header_map[str(name).strip()] = idx
        return header_map
        
    def ensure_column(self, col_display_name: str) -> int:
        """
        确保指定的显示列名存在于首行表头中；若不存在则追加并返回列索引（1-based）。
        """
        col_idx = self.header_map.get(col_display_name)
        if col_idx:
            return col_idx
        # 计算新的列索引：在现有最大索引之后
        max_idx = max(self.header_map.values()) if self.header_map else 0
        col_idx = max_idx + 1
        # 在首行写入新的表头
        self.ws.cell(row=1, column=col_idx, value=col_display_name)
        # 更新映射并记录日志
        self.header_map[col_display_name] = col_idx
        logger.info(f"excel_header_column_created name={col_display_name} index={col_idx}")
        return col_idx

    def iter_rows(self) -> Tuple[int, Dict[str, Any]]:
        # 从第二行开始（跳过表头）
        for i, row in enumerate(self.ws.iter_rows(min_row=2, values_only=False), start=2):
            yield i, { "cells": row }

    def get_value(self, row_cells, col_display_name: str) -> Optional[str]:
        """
        读取指定列的值；若列不存在返回 None。通过工作表坐标读取，避免 row_cells 不包含新增列的情况。
        """
        col_idx = self.header_map.get(col_display_name)
        if not col_idx:
            return None
        # 通过行号定位
        row_num = row_cells[0].row if row_cells else 1
        v = self.ws.cell(row=row_num, column=col_idx).value
        if v is None:
            return None
        return str(v).strip()

    def set_value(self, row_cells, col_display_name: str, value: Any) -> None:
        """
        写入指定列的值；若列不存在则自动新增到表头末尾后再写入。
        通过工作表坐标写入，避免 row_cells 不包含新增列的情况。
        """
        col_idx = self.header_map.get(col_display_name)
        if not col_idx:
            col_idx = self.ensure_column(col_display_name)
        row_num = row_cells[0].row if row_cells else self.ws.max_row
        self.ws.cell(row=row_num, column=col_idx, value=value)

    def save(self) -> None:
        self.wb.save(self.excel_path)