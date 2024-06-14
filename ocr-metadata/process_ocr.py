#函数化，图片+json提取，指定文件夹下图片依次读取执行
# v0.5.1，结果保存到本地excel中
# v0.5.2，将所有控制台的print内容保存到本地logs.txt文件中
# v0.5.3，json输出元数据结果

import base64
import requests
import os
import random
import time
from openai import OpenAI
import openpyxl
import logging
import os
from dotenv import load_dotenv

load_dotenv()


def image_to_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def image_to_base64_url(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8") 

def split_ocr_text(ocr_result):
    header = []
    title_left = []
    title_right = []
    content_section = []
    footer = []

    for item in ocr_result['words_result']:
        words = item['words']
        location = item['location']
        # 第一段:页头
        if location['top'] < 350:
            header.append(words)
        # 第二段:标题部分
        elif 350 < location['top'] < 1300:
            if location['left'] < 1000:  # 根据左右坐标分割
                title_left.append(words)
            else:
                title_right.append(words)
        # 第三段:内容部分
        elif 1300 < location['top'] < 3000: 
            content_section.append(words)
        # 第四段:页脚
        else:
            footer.append(words)

    return {
        'title_left': ' '.join(title_left),
        'title_right': ' '.join(title_right),
        'content_section': ' '.join(content_section),
        'footer': ' '.join(footer)
    }

def get_ocr_result(image_path, ocr_url, access_token):
    ocr_img = image_to_base64(image_path)
    ocr_url = ocr_url
    params = {"image": ocr_img}
    request_url = ocr_url + "?access_token=" + access_token
    headers = {'content-type': 'application/x-www-form-urlencoded'}
    response = requests.post(request_url, data=params, headers=headers)
    if response:
        # return response.json()
        ocr_result = response.json()
        logging.info(f"OCR请求成功: {image_path}")
        logging.info(f"OCR结果: {ocr_result}")
        return ocr_result
    else:
        logging.error(f"OCR请求失败: {image_path}")
        return None

def llm_ocr_metadata(ocr_result, image_path, base_url, api_key):

    base64_image = image_to_base64_url(image_path)

    url = base_url
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    prompt = """
        ## 角色与技能
        - 你是一名图书馆高级元数据馆员。
        - 从OCR生成的json数据中，提取信息创建符合要求的资源描述元数据。
        - json数据包含了文本及其在图片中的坐标信息。这些坐标信息与“图片布局说明”以及图片互为参照，可以作为你生成元数据的参考。

        ## 图片布局说明
        图片中从上到下的布局分成3段，分别是：
        1. 第一段：
        - 左侧加粗字体是文档 `标题`。
        - 右侧小字体分别是：`分发类别`， `编号`， `时间`， `语种说明`。
        - **第一段与第二段之间有加粗分隔线**

        2. 第二段：
        - 偏左侧的小字体是 `目录`。
        - 随后是正文的题目及其内容。
        - **左侧的小字体 `目录` 与正文内容之间在图片布局中有空白间隔**

        3. 第三段：
        - 文档的 `馆藏条码`、`GE编号`。
        - `馆藏条码`文本与条形码在一起，以`WBDA`开头。
        - `GE编号`一般在图片左下角，字体较小。以`GE`或`数字`开头的编号
        - 注意：`馆藏条码`信息也有可能在 `第一段`中

        ## 输出要求
        请根据上述的图片布局说明和用户提供的含有OCR文本及其坐标位置的Json数据，以及读取的图片，综合判断。严格按照如下要求提取信息：
        标题、分发类别、编号、时间、语种说明、目录、馆藏条码、GE编号

        ## 输出格式与样例：
        - 使用json格式输出。输出样例如下：
        - 输出样例1：
        {
            "标题": "联合国贸易和发展会议",
            "分发类别": "Distr. GENERAL",
            "编号": "TD/B/1215",
            "时间": "11 May 1989",
            "语种说明": "CHINESE Original:SPANISH",
            "目录": "目录1 -- 目录2 -- 目录3",
            "馆藏条码": "WBDA015253",
            "GE编号": "GE.89-51367"
        }
        - 输出样例2：
        {
            "标题": "UNITED NATIONS Security Council",
            "分发类别": "Distr. GENERAL",
            "编号": "S/16933",
            "时间": "5 February 1985",
            "语种说明": "ORIGINAL: ENGLISH",
            "目录": "",
            "馆藏条码": "WBDA015590",
            "GE编号": "85-03348 1558u (E)"
        }

        ## 限制
        - 严格按照输出的字段要求提取信息。
        - 不要输出任何多余的信息。
        - 不要有任何的解释或者说明。
        - 只能依据提供的文本提取信息，并保持一致，不能捏造。
        - 请不要联网搜索。

        按照如上要求，请一步步思考。

        """
    data = {
        "model": "gpt-4o",
        "messages": [
            {"role": "system", "content": prompt},
            {
            "role": "user",
            "content": [
                {
                "type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}
                },
                {"type": "text", "text": str(ocr_result)}
            ]
            }
        ],
        }

    try:
        response = requests.post(url, headers=headers, json=data)
        # print(response.text)
        if response.status_code == 200:
            choices = response.json().get('choices', [])
            if choices:
                ocr_metadata_output = choices[0]['message']['content']
                print(f"LLM-OCR元数据:\n\n {ocr_metadata_output}")
                logging.info(f"LLM-OCR元数据:\n\n {ocr_metadata_output}")
                return ocr_metadata_output
            else:
                print("No choices found in the response.")
                logging.info("No choices found in the response.")
        else:
            print("Response Code:", response.status_code)
            logging.error(f"Response Code: {response.status_code}")
    except Exception as e:
        print("Error:", e)
        logging.error(f"Error: {e}")


def llm_formatting(formatted_output, ocr_result):
    api_key = os.getenv("BAICHUAN_API_KEY")
    base_url = os.getenv("BAICHUAN_BASE_URL")

    if not api_key or not base_url:
        print("百川API配置未设置")
        logging.error("百川API配置未设置")
        return None

    client = OpenAI(api_key=api_key, base_url=base_url)
    
    prompt = """
        ## 角色与技能
        - 文本分析与排版助手。
        - 用户输入的原始 OCR 文本按照固定坐标格式化整理，分成了三大类。但由于排版的多样性，根据坐标固定分类有可能错误，你可以根据下列的“图片布局说明”和含有坐标位置的OCR生成Json数据，对格式化文本内容重新整理分类，确保准确性。

        ## 图片布局说明
        图片中从上到下的布局一般分成3段，分别是：
        1. 第一段：
        - 左侧加粗字体是文档 `标题`。
        - 右侧小字体分别是：`分发类别`， `编号`， `时间`， `语种说明`。
        - **第一段与第二段之间有加粗分隔线**

        2. 第二段：
        - 偏左侧的小字体是 `目录`。
        - 随后是正文的题目及其内容。
        - **左侧的小字体 `目录` 与正文内容之间在图片布局中有空白间隔**

        3. 第三段：
        - 文档的 `馆藏条码`、`GE编号`。
        - `馆藏条码`文本与条形码在一起，以`WBDA`开头。
        - `GE编号`一般在图片左下角，字体较小。以`GE`或`数字`开头的编号
        - 注意：`馆藏条码`信息也有可能在 `第一段`中

        ## 输出要求
        - 请根据上述的“图片布局说明”和用户提供的含有OCR文本及其坐标位置的Json数据，对格式化文本内容重新整理。
        - 使用 markdown 格式
        - 严格按照如下要求提取信息：

        ```
        - 元数据信息
            - 标题：
            - 目录：
            - 编号：
            - 时间：
            - 语种说明：
            - 馆藏条码：
            - GE编号：
        - 内容

        ```

        ## 限制
        - 严格按照输出的字段要求提取信息。
        - 不要输出任何多余的信息。
        - 不要有任何的解释、说明或者翻译。
        - 只能依据提供的文本提取信息，并保持一致，不能捏造。
        - 请不要联网搜索。

        按照如上要求，请一步步思考，确保结果严格按照要求输出。

        """
    try:
        response = client.chat.completions.create(
            model="Baichuan4",
            temperature=0.0,
            messages=[
                {"role": "system", "content": prompt},
                {"role": "user", "content": f"已格式化的输出:\n{formatted_output}\n\nOCR原始结果:\n {ocr_result}"},
            ],
            stream=False
        )
    except Exception as e:
        print(f"API请求失败: {e}")
        logging.error(f"API请求失败: {e}")
        return None

    if response and response.choices:
        llm_formatting_output = response.choices[0].message.content
        print(f"LLM-OCR原文整理:\n\n {llm_formatting_output}")
        logging.info(f"LLM-OCR原文整理:\n\n {llm_formatting_output}")
        return llm_formatting_output
    else:
        print("API响应无效或没有选项")
        logging.error("API响应无效或没有选项")
        return None
    
def save_to_excel(folder_path, file_name, image_name, formatted_output, llm_formatting_output, ocr_metadata_output):
    output_path = os.path.join(folder_path, file_name)
    
    # 检查文件是否存在
    if os.path.exists(output_path):
        workbook = openpyxl.load_workbook(output_path)
    else:
        workbook = openpyxl.Workbook()
    
    # 获取活动工作表
    worksheet = workbook.active
    
    # 检查是否需要写入列名
    if worksheet.max_row == 1:
        worksheet['A1'] = '图片名称'
        worksheet['B1'] = 'OCR原文'
        worksheet['C1'] = '格式化输出'
        worksheet['D1'] = 'OCR元数据'
    
    # 写入数据
    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1, value=image_name)
    worksheet.cell(row=row, column=2, value=formatted_output)
    worksheet.cell(row=row, column=3, value=llm_formatting_output)
    worksheet.cell(row=row, column=4, value=ocr_metadata_output)
    
    # 保存 Excel 文件
    workbook.save(output_path)
    print(f"Excel 文件已保存: {output_path}")
    logging.info(f"Excel 文件已保存: {output_path}")    

def main(folder_path):
    # 从环境变量中读取OCR配置
    ocr_url = os.getenv("OCR_URL")
    access_token = os.getenv("OCR_ACCESS_TOKEN")

    if not ocr_url or not access_token:
        print("OCR配置未设置")
        logging.error("OCR配置未设置")
        return

    # 从环境变量中读取GPT-4o配置
    base_url = os.getenv("GPT4O_BASE_URL")
    api_key = os.getenv("GPT4O_API_KEY")

    if not base_url or not api_key:
        print("GPT-4o配置未设置")
        logging.error("GPT-4o配置未设置")
        return
    
    image_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(('.jpg', '.png'))]
    for image_path in image_files:
        image_name = os.path.basename(image_path)
        print(f"当前处理图片名: {image_name}")
        logging.info(f"当前处理图片名: {image_name}")
        ocr_result = get_ocr_result(image_path, ocr_url, access_token)
        if ocr_result:
            separated_text = split_ocr_text(ocr_result)
            formatted_output = f"""
---

## 标题、目录、编号、年月与语种
左侧: {separated_text['title_left']}
右侧: {separated_text['title_right']}

## 内容
{separated_text['content_section']}

## 馆藏条码与GE编号
{separated_text['footer']}

---
"""
            print(f"OCR原文: {formatted_output}")
            logging.info(f"OCR原文: {formatted_output}")
            llm_formatting_output = llm_formatting(formatted_output, ocr_result)
            ocr_metadata_output = llm_ocr_metadata(ocr_result, image_path, base_url, api_key)
            # 保存结果到 Excel 文件
            save_to_excel(folder_path, 'output.xlsx', image_name, formatted_output, llm_formatting_output, ocr_metadata_output)
        else:
            print(f"OCR请求失败: {image_path}")
            logging.info(f"OCR请求失败: {image_path}")
        time.sleep(random.randint(1, 3))

if __name__ == "__main__":
    FOLDER_PATH = r'E:\scripts\AI-demo\1' # 替换为你的文件夹路径

    # 保存日志文件
    log_file = os.path.join(FOLDER_PATH, 'logs.txt')

    # 如果日志文件已存在,设置filemode='a'以追加写入
    if os.path.exists(log_file):
        filemode = 'a'
    else:
        filemode = 'w'

    logging.basicConfig(
        filename=log_file,
        filemode=filemode, # 'a'为追加模式,'w'为覆盖模式
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    main(FOLDER_PATH)