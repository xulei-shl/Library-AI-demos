"""
候选图书 LLM 筛选器
负责读取 Excel 中的候选图书，分组调用 LLM 筛选，并将结果写回 Excel
"""

import json
import random
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook

from src.utils.logger import get_logger
from src.utils.llm.client import UnifiedLLMClient

logger = get_logger(__name__)


class CandidateFilter:
    """候选图书 LLM 筛选器"""

    def __init__(self, config: dict):
        """
        Args:
            config: literature_fm.yaml 中的 candidate_filter 配置节
        """
        self.config = config
        self.llm_client = UnifiedLLMClient()
        self.excluded_sheets = config.get('excluded_sheets', ['主题汇总', '元信息'])
        self.group_config = config.get('grouping', {})
        self.group_size = self.group_config.get('group_size', 6)
        self.excel_fields = config.get('excel_fields', {})

    def filter_from_excel(self, excel_path: str) -> dict:
        """
        从 Excel 文件进行候选图书筛选

        Args:
            excel_path: Excel 文件路径（包含主题汇总和多个候选图书sheet）

        Returns:
            {
                'success': True/False,
                'processed': 处理总数,
                'passed': 通过筛选数,
                'groups': 分组数量,
                'output_file': '原文件路径（已修改）',
                'error': '错误信息'
            }
        """
        try:
            # 1. 读取主题信息
            theme_info = self._read_theme_info(excel_path)
            if not theme_info:
                return {
                    'success': False,
                    'error': '未能读取到主题汇总信息，请确认 Excel 中存在「主题汇总」sheet'
                }

            logger.info(f"✓ 读取主题信息: {theme_info.get('theme_name', '')}")

            # 2. 读取候选图书 sheets
            candidate_sheets = self._read_candidate_sheets(excel_path)
            if not candidate_sheets:
                return {
                    'success': False,
                    'error': '未找到候选图书 sheet（所有 sheet 都被排除）'
                }

            logger.info(f"✓ 读取到 {len(candidate_sheets)} 个候选图书 sheet")

            # 3. 合并所有候选图书
            all_candidates = []
            for sheet_info in candidate_sheets:
                sheet_name = sheet_info['sheet_name']
                df = sheet_info['df']
                for _, row in df.iterrows():
                    book_data = {
                        'sheet_name': sheet_name,
                        'index': row.name  # 记录原始行索引
                    }
                    for field in self.excel_fields.get('book_fields', []):
                        book_data[field] = row.get(field, '')
                    all_candidates.append(book_data)

            total_count = len(all_candidates)
            logger.info(f"✓ 合并共 {total_count} 本候选图书")

            # 4. 分组
            groups = self._group_candidates(all_candidates)
            logger.info(f"✓ 分成 {len(groups)} 组进行筛选")

            # 5. 分组调用 LLM 筛选
            filtered_candidates = []
            for idx, group in enumerate(groups, 1):
                logger.info(f"{'='*60}")
                logger.info(f"处理第 {idx}/{len(groups)} 组 (共 {len(group)} 本)")
                logger.info(f"{'='*60}")

                filtered_group = self._call_llm_filter(theme_info, group, idx)
                filtered_candidates.extend(filtered_group)

            # 6. 按原始 sheet 重新组织结果
            sheet_results = self._organize_by_sheet(filtered_candidates)

            # 7. 写入 Excel
            self._write_results_to_excel(excel_path, sheet_results)

            # 统计通过筛选数量
            passed_count = sum(1 for book in filtered_candidates if book.get('初评结果') == '通过')

            return {
                'success': True,
                'processed': total_count,
                'passed': passed_count,
                'groups': len(groups),
                'output_file': excel_path
            }

        except Exception as e:
            logger.error(f"筛选过程异常: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }

    def _read_theme_info(self, excel_path: str) -> dict:
        """
        从 '主题汇总' sheet 提取主题信息

        Returns:
            {
                'theme_name': 'xxx',
                'slogan': 'xxx',
                'description': 'xxx',
                'target_vibe': 'xxx'
            }
        """
        try:
            df = pd.read_excel(excel_path, sheet_name='主题汇总')

            if df.empty:
                return {}

            # 取第一行数据
            row = df.iloc[0]

            field_mapping = self.excel_fields.get('theme_summary', {})

            return {
                'theme_name': row.get(field_mapping.get('theme_name', '主题名称'), ''),
                'slogan': row.get(field_mapping.get('slogan', '副标题'), ''),
                'description': row.get(field_mapping.get('description', '情境描述'), ''),
                'target_vibe': row.get(field_mapping.get('target_vibe', '预期氛围'), '')
            }

        except Exception as e:
            logger.warning(f"读取主题汇总失败: {e}")
            return {}

    def _read_candidate_sheets(self, excel_path: str) -> List[Dict]:
        """
        读取候选图书 sheet（排除配置的 sheet）

        Returns:
            List[Dict]: [
                {
                    'sheet_name': 'sheet名',
                    'df': DataFrame
                }
            ]
        """
        try:
            xls = pd.ExcelFile(excel_path)
            all_sheets = xls.sheet_names

            # 过滤排除的sheet
            candidate_sheets = [s for s in all_sheets if s not in self.excluded_sheets]

            results = []
            for sheet_name in candidate_sheets:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                results.append({
                    'sheet_name': sheet_name,
                    'df': df
                })

            return results

        except Exception as e:
            logger.error(f"读取候选 sheets 失败: {e}", exc_info=True)
            return []

    def _group_candidates(self, candidates: List[Dict]) -> List[List[Dict]]:
        """
        随机分组候选图书（每组约6条）

        Args:
            candidates: 所有候选图书列表

        Returns:
            List[List[Dict]]: 分组后的列表
        """
        shuffle = self.group_config.get('shuffle', True)
        random_seed = self.group_config.get('random_seed')

        # 设置随机种子（便于复现）
        if random_seed is not None:
            random.seed(random_seed)

        # 复制并打乱
        if shuffle:
            candidates = candidates.copy()
            random.shuffle(candidates)

        # 分组
        groups = []
        for i in range(0, len(candidates), self.group_size):
            groups.append(candidates[i:i + self.group_size])

        return groups

    def _call_llm_filter(self, theme: dict, group: List[Dict], group_idx: int) -> List[Dict]:
        """
        调用 LLM 筛选单组图书

        Args:
            theme: 主题信息字典
            group: 单组候选图书
            group_idx: 组索引（用于日志）

        Returns:
            List[Dict]: 带筛选结果的图书列表
                每个元素新增: 初评结果, 初评分数, 初评理由
        """
        try:
            # 构建用户提示词
            user_prompt = self._build_user_prompt(theme, group)

            # 调用 LLM
            task_name = self.config.get('llm', {}).get('task_name', 'literary_candidate_filter')
            response = self.llm_client.call(
                task_name=task_name,
                user_prompt=user_prompt
            )

            # 解析响应
            results = self._parse_llm_response(response)

            # 将结果合并回原数据（使用中文字段名）
            for book, result in zip(group, results):
                is_pass = result.get('is_pass', False)
                book['初评结果'] = '通过' if is_pass else '未通过'
                book['初评分数'] = result.get('score', 0)
                book['初评理由'] = result.get('reason', '')

            logger.info(f"  ✓ 第 {group_idx} 组筛选完成")
            return group

        except Exception as e:
            logger.error(f"  ✗ 第 {group_idx} 组筛选失败: {e}", exc_info=True)
            # 失败时标记为未通过
            for book in group:
                book['初评结果'] = '未通过'
                book['初评分数'] = 0
                book['初评理由'] = f"筛选失败: {str(e)}"
            return group

    def _build_user_prompt(self, theme: dict, books: List[Dict]) -> str:
        """
        构建用户提示词（主题信息 + 图书列表）
        """
        # 主题信息
        theme_part = f"""# Target Theme（策划主题）
主题名称：{theme.get('theme_name', '')}
副标题：{theme.get('slogan', '')}
情境描述：{theme.get('description', '')}
预期氛围：{theme.get('target_vibe', '')}
"""

        # 图书列表（JSON格式）
        book_fields = self.excel_fields.get('book_fields', [])
        books_json = []
        for idx, book in enumerate(books, 1):
            book_data = {"序号": idx}
            for field in book_fields:
                book_data[field] = book.get(field, '')
            books_json.append(book_data)

        books_part = "\n# Candidate Books（候选图书）\n" + json.dumps(
            books_json, ensure_ascii=False, indent=2
        )

        return theme_part + "\n" + books_part + "\n\n请按照要求的JSON格式输出筛选结果。"

    def _parse_llm_response(self, response: str) -> List[Dict]:
        """
        解析 LLM 响应

        Args:
            response: LLM 返回的 JSON 字符串

        Returns:
            List[Dict]: 解析后的结果列表
        """
        try:
            # 尝试直接解析 JSON
            results = json.loads(response)

            # 如果是单个对象，包装成列表
            if isinstance(results, dict):
                results = [results]

            return results

        except json.JSONDecodeError:
            # 尝试提取 JSON 代码块
            import re
            json_match = re.search(r'```(?:json)?\s*(\[.*?\]|\{.*?\})\s*```', response, re.DOTALL)
            if json_match:
                try:
                    results = json.loads(json_match.group(1))
                    if isinstance(results, dict):
                        results = [results]
                    return results
                except json.JSONDecodeError:
                    pass

            logger.error(f"JSON 解析失败，原始响应: {response[:500]}...")
            # 返回空结果
            return []

    def _organize_by_sheet(self, candidates: List[Dict]) -> List[Dict]:
        """
        按原始 sheet 重新组织结果

        Args:
            candidates: 所有候选图书（带筛选结果）

        Returns:
            List[Dict]: [
                {
                    'sheet_name': 'xxx',
                    'df_with_results': DataFrame
                }
            ]
        """
        # 按 sheet_name 分组
        sheet_groups = {}
        for book in candidates:
            sheet_name = book['sheet_name']
            if sheet_name not in sheet_groups:
                sheet_groups[sheet_name] = []
            sheet_groups[sheet_name].append(book)

        # 构建结果
        results = []
        for sheet_name, books in sheet_groups.items():
            # 按原始行索引排序
            books_sorted = sorted(books, key=lambda x: x['index'])

            # 构建 DataFrame
            df_data = []
            for book in books_sorted:
                row = {k: v for k, v in book.items() if k not in ['sheet_name', 'index']}
                df_data.append(row)

            df = pd.DataFrame(df_data)
            results.append({
                'sheet_name': sheet_name,
                'df_with_results': df
            })

        return results

    def _write_results_to_excel(self, excel_path: str, sheet_results: List[Dict]) -> str:
        """
        将筛选结果写入 Excel 新增列

        Args:
            excel_path: 原文件路径
            sheet_results: 各sheet的筛选结果

        Returns:
            str: 输出文件路径（原文件）
        """
        try:
            # 加载现有工作簿
            book = load_workbook(excel_path)

            new_cols = ['初评结果', '初评分数', '初评理由', '人工评选', '人工推荐语', '终评理由']

            for sheet_result in sheet_results:
                sheet_name = sheet_result['sheet_name']
                df = sheet_result['df_with_results']

                if sheet_name not in book.sheetnames:
                    logger.warning(f"Sheet 不存在，跳过: {sheet_name}")
                    continue

                ws = book[sheet_name]

                # 获取现有列数
                existing_cols = ws.max_column

                # 写入表头
                for col_idx, col_name in enumerate(new_cols, 1):
                    ws.cell(row=1, column=existing_cols + col_idx, value=col_name)

                # 写入数据
                for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                    for col_idx, col_name in enumerate(new_cols, 1):
                        value = row_data.get(col_name, '')
                        ws.cell(row=row_idx, column=existing_cols + col_idx, value=value)

                logger.info(f"  ✓ 写入 sheet: {sheet_name}")

            # 保存
            book.save(excel_path)
            book.close()

            logger.info(f"✓ 结果已写入: {excel_path}")
            return excel_path

        except Exception as e:
            logger.error(f"写入 Excel 失败: {e}", exc_info=True)
            raise
