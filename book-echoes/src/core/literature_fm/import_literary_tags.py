"""
导入 xlsx/csv 数据到 literary_tags 表的工具
用法:
    python F:/Github/Library-AI-demos/book-echoes/src/core/literature_fm/import_literary_tags.py <file_path> [--db <db_path>]
"""

import argparse
import csv
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


DB_DEFAULT = "runtime/database/books_history.db"
TABLE_NAME = "literary_tags"


def get_db_connection(db_path: str):
    """获取数据库连接"""
    db_file = Path(db_path)
    db_file.parent.mkdir(parents=True, exist_ok=True)
    return sqlite3.connect(db_path)


def read_xlsx(file_path: str) -> list[dict]:
    """读取 xlsx 文件"""
    if openpyxl is None:
        raise ImportError("需要安装 openpyxl 库: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 获取表头
    headers = [cell.value for cell in ws[1]]

    # 读取数据
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if header is not None:
                row_dict[header] = row[i]
        if row_dict:
            rows.append(row_dict)

    return rows


def read_csv(file_path: str) -> list[dict]:
    """读取 csv 文件"""
    rows = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 过滤空行
            if any(v for v in row.values()):
                rows.append(row)
    return rows


def read_file(file_path: str) -> list[dict]:
    """根据文件扩展名读取文件"""
    path = Path(file_path)
    if path.suffix.lower() == '.xlsx':
        return read_xlsx(file_path)
    elif path.suffix.lower() == '.csv':
        return read_csv(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {path.suffix}")


def import_data(db_path: str, rows: list[dict], replace: bool = False) -> tuple[int, int]:
    """
    导入数据到数据库

    Args:
        db_path: 数据库路径
        rows: 数据行列表
        replace: 是否替换已存在的记录

    Returns:
        (成功数量, 失败数量)
    """
    conn = get_db_connection(db_path)
    cursor = conn.cursor()

    columns = [
        'book_id', 'call_no', 'title', 'tags_json',
        'llm_model', 'llm_provider', 'llm_status', 'retry_count', 'error_message'
    ]

    insert_sql = f"""
        INSERT INTO {TABLE_NAME} ({', '.join(columns)})
        VALUES ({', '.join(['?' for _ in columns])})
    """

    if replace:
        insert_sql += " ON CONFLICT(book_id) DO UPDATE SET "
        update_parts = [f"{col} = excluded.{col}" for col in columns if col != 'book_id']
        insert_sql += ", ".join(update_parts)
        insert_sql += ", updated_at = CURRENT_TIMESTAMP"

    success = 0
    failed = 0

    for i, row in enumerate(rows, 1):
        try:
            values = []
            for col in columns:
                val = row.get(col)

                # 类型转换
                if col == 'retry_count':
                    val = int(val) if val is not None else 0
                elif val == '' or val is None:
                    val = None

                values.append(val)

            cursor.execute(insert_sql, values)
            success += 1

        except Exception as e:
            print(f"  ✗ 第 {i} 行失败: {e}")
            failed += 1

    conn.commit()
    conn.close()

    return success, failed


def show_table_structure():
    """显示表结构"""
    print(f"\n{TABLE_NAME} 表结构:")
    print("-" * 60)
    print(f"{'列名':<20} {'类型':<20} {'说明'}")
    print("-" * 60)
    structure = [
        ('id', 'INTEGER PRIMARY KEY', '自增ID'),
        ('book_id', 'INTEGER NOT NULL', '书籍ID（唯一约束）'),
        ('call_no', 'TEXT', '索书号'),
        ('title', 'TEXT', '标题'),
        ('tags_json', 'TEXT', '标签数据（JSON格式）'),
        ('llm_model', 'TEXT', 'LLM模型'),
        ('llm_provider', 'TEXT', 'LLM提供者'),
        ('llm_status', 'TEXT DEFAULT pending', 'LLM状态'),
        ('retry_count', 'INTEGER DEFAULT 0', '重试次数'),
        ('error_message', 'TEXT', '错误信息'),
        ('created_at', 'TIMESTAMP', '创建时间'),
        ('updated_at', 'TIMESTAMP', '更新时间'),
    ]
    for col, typ, desc in structure:
        print(f"{col:<20} {typ:<20} {desc}")
    print("-" * 60)


def main():
    parser = argparse.ArgumentParser(
        description=f"导入 xlsx/csv 数据到 {TABLE_NAME} 表",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    python import_literary_tags.py data.xlsx
    python import_literary_tags.py data.csv --db runtime/database/books_history.db
    python import_literary_tags.py data.xlsx --replace
        """
    )

    parser.add_argument('file', help='要导入的 xlsx 或 csv 文件路径')
    parser.add_argument('--db', default=DB_DEFAULT, help=f'数据库路径 (默认: {DB_DEFAULT})')
    parser.add_argument('--replace', action='store_true', help='替换已存在的记录（按book_id）')
    parser.add_argument('--show-structure', action='store_true', help='显示表结构')

    args = parser.parse_args()

    if args.show_structure:
        show_table_structure()
        return

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"✗ 文件不存在: {file_path}")
        sys.exit(1)

    # 读取文件
    print(f"\n正在读取文件: {file_path}")
    try:
        rows = read_file(str(file_path))
        print(f"✓ 成功读取 {len(rows)} 条数据")
    except Exception as e:
        print(f"✗ 读取文件失败: {e}")
        sys.exit(1)

    # 显示列名
    if rows:
        print(f"文件列名: {', '.join(rows[0].keys())}")

    # 导入数据
    print(f"\n正在导入到数据库: {args.db}")
    success, failed = import_data(args.db, rows, args.replace)

    print(f"\n导入完成:")
    print(f"  ✓ 成功: {success}")
    print(f"  ✗ 失败: {failed}")


if __name__ == "__main__":
    main()
