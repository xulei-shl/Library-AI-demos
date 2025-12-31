"""
结果导出器
导出主题书架结果到 Excel
"""

import json
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

from src.utils.logger import get_logger

logger = get_logger(__name__)

# books表中需要导出的douban_字段列表
DOUBAN_FIELDS = [
    'barcode',
    'douban_url', 'douban_rating', 'douban_title', 'douban_subtitle',
    'douban_original_title', 'douban_author', 'douban_translator',
    'douban_publisher', 'douban_producer', 'douban_series',
    'douban_series_link', 'douban_price', 'douban_isbn',
    'douban_pages', 'douban_binding', 'douban_pub_year',
    'douban_rating_count', 'douban_summary', 'douban_author_intro',
    'douban_catalog', 'douban_cover_image'
]


class ThemeExporter:
    """导出主题书架结果"""

    def __init__(self, output_dir: str = "runtime/outputs/theme_shelf", db_path: str = "runtime/database/books_history.db"):
        """
        初始化导出器

        Args:
            output_dir: 输出目录
            db_path: 数据库路径
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.db_path = db_path

    def _fetch_books_details(self, book_ids: List[int]) -> Dict[int, Dict]:
        """
        从books表查询书籍详细信息

        Args:
            book_ids: 书籍ID列表

        Returns:
            Dict[int, Dict]: book_id -> 详细信息字典
        """
        if not book_ids:
            return {}

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row

            # 构建查询字段
            fields = ['id'] + DOUBAN_FIELDS
            query = f"SELECT {', '.join(fields)} FROM books WHERE id IN ({','.join('?' * len(book_ids))})"

            cursor = conn.execute(query, book_ids)
            details = {}
            for row in cursor.fetchall():
                details[row['id']] = {field: row[field] for field in fields if field != 'id'}

            conn.close()
            return details

        except Exception as e:
            logger.warning(f"查询books表详细信息失败: {str(e)}")
            return {}

    def export_excel(
        self,
        results: List[Dict],
        theme_text: str,
        conditions: Dict,
        output_path: Optional[str] = None
    ) -> str:
        """
        导出 Excel 格式的主题书架

        Args:
            results: 融合后的结果列表
            theme_text: 用户输入的主题
            conditions: 解析后的检索条件
            output_path: 自定义输出路径

        Returns:
            实际输出的文件路径
        """
        # 1. 从books表查询详细信息
        book_ids = [item.get('book_id') for item in results if item.get('book_id')]
        books_details = self._fetch_books_details(book_ids)

        # 2. 构建 DataFrame
        data = []

        for i, item in enumerate(results, 1):
            tags = self._parse_tags(item.get('tags_json', ''))
            match_tags = self._get_match_tags(conditions, tags)

            # 基础字段
            row_data = {
                '序号': i,
                '书名': item.get('title', ''),
                '作者': item.get('author', ''),
                '索书号': item.get('call_no', ''),
                '综合评分': round(item.get('final_score', 0), 3),
                '向量得分': round(item.get('vector_score', 0), 3),
                '标签得分': round(item.get('tag_score', 0), 3),
                '匹配标签': '、'.join(match_tags),
                '来源': self._get_source_label(item)
            }

            # 添加books表的额外字段
            book_id = item.get('book_id')
            if book_id and book_id in books_details:
                detail = books_details[book_id]
                for field in DOUBAN_FIELDS:
                    # 字段名映射：douban_xxx -> 豆瓣_xxx，barcode -> 条码
                    if field == 'barcode':
                        row_data['条码'] = detail.get(field, '')
                    else:
                        row_data[field] = detail.get(field, '')

            data.append(row_data)

        df = pd.DataFrame(data)

        # 2. 生成输出文件名
        if output_path:
            output_file = Path(output_path)
        else:
            output_file = self._generate_filename(theme_text)

        # 确保父目录存在
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # 3. 导出 Excel（使用 openpyxl）
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 主结果表
                df.to_excel(writer, sheet_name='推荐结果', index=False)

                # 设置列宽
                from openpyxl.utils import get_column_letter
                worksheet = writer.sheets['推荐结果']

                # 基础列宽度
                base_widths = [5, 35, 15, 12, 10, 10, 10, 30, 8, 15]  # 前基础列+条码列

                # douban字段列宽度
                douban_widths = {
                    'douban_url': 40,
                    'douban_title': 30,
                    'douban_subtitle': 25,
                    'douban_original_title': 30,
                    'douban_author': 20,
                    'douban_translator': 20,
                    'douban_publisher': 20,
                    'douban_producer': 20,
                    'douban_series': 25,
                    'douban_series_link': 40,
                    'douban_price': 10,
                    'douban_isbn': 15,
                    'douban_pages': 10,
                    'douban_binding': 12,
                    'douban_pub_year': 12,
                    'douban_rating': 10,
                    'douban_rating_count': 12,
                    'douban_summary': 50,
                    'douban_author_intro': 40,
                    'douban_catalog': 40,
                    'douban_cover_image': 40
                }

                # 按DOUBAN_FIELDS顺序构建完整列宽列表
                all_widths = base_widths[:]
                for field in DOUBAN_FIELDS:
                    if field == 'barcode':
                        continue  # 已经在base_widths中
                    all_widths.append(douban_widths.get(field, 15))

                # 应用列宽
                for i, width in enumerate(all_widths, 1):
                    col_letter = get_column_letter(i)
                    worksheet.column_dimensions[col_letter].width = width

                # 元信息表
                meta_data = [
                    ['主题', theme_text],
                    ['检索条件', str(conditions)],
                    ['推荐时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                    ['推荐数量', len(results)]
                ]
                meta_df = pd.DataFrame(meta_data, columns=['项目', '内容'])
                meta_df.to_excel(writer, sheet_name='元信息', index=False)

            logger.info(f"导出成功: {output_file}")
            return str(output_file)

        except Exception as e:
            logger.error(f"导出失败: {str(e)}")
            raise

    def _parse_tags(self, tags_json: str) -> Dict:
        """解析标签 JSON"""
        try:
            return json.loads(tags_json) if tags_json else {}
        except (json.JSONDecodeError, TypeError):
            return {}

    def _get_match_tags(self, conditions: Dict, tags: Dict) -> List[str]:
        """获取匹配的标签"""
        matched = []
        for dim, dim_tags in conditions.items():
            if isinstance(dim_tags, list):
                for tag in dim_tags:
                    book_tags = tags.get(dim, [])
                    if isinstance(book_tags, list) and tag in book_tags:
                        matched.append(tag)
        return matched

    def _get_source_label(self, item: Dict) -> str:
        """获取来源标签"""
        sources = item.get('sources', [])
        if 'vector' in sources and 'tag' in sources:
            return '混合'
        elif 'vector' in sources:
            return '向量'
        elif 'tag' in sources:
            return '标签'
        return '未知'

    def _generate_filename(self, theme_text: str) -> Path:
        """生成输出文件名"""
        # 清理文件名中的非法字符
        safe_name = re.sub(r'[^\w\s-]', '', theme_text)[:15]
        safe_name = re.sub(r'[\s-]+', '_', safe_name).strip('_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return self.output_dir / f"主题书架_{safe_name}_{timestamp}.xlsx"

    def export_theme_batch(
        self,
        theme_results: List[Dict],
        output_path: Optional[str] = None
    ) -> str:
        """
        导出批量主题检索结果（新格式：混合检索 + RRF融合）

        Args:
            theme_results: 主题结果列表，格式：
                [
                    {
                        "theme": {...},  # original_theme
                        "books": [...],
                        "filter_conditions": [...],
                        "search_keywords": [...],
                        "stats": {...}
                    }
                ]
            output_path: 自定义输出路径

        Returns:
            实际输出的文件路径
        """
        # 1. 收集所有书籍ID
        all_book_ids = []
        for theme_result in theme_results:
            books = theme_result.get('books', [])
            for book in books:
                book_id = book.get('book_id')
                if book_id and book_id not in all_book_ids:
                    all_book_ids.append(book_id)

        # 2. 查询详细信息
        books_details = self._fetch_books_details(all_book_ids)

        # 3. 构建 DataFrame（每个主题一个sheet）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        if output_path:
            output_file = Path(output_path)
        else:
            output_file = self.output_dir / f"主题书架_批量_{timestamp}.xlsx"

        output_file.parent.mkdir(parents=True, exist_ok=True)

        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 1. 汇总表
                summary_data = []
                for idx, theme_result in enumerate(theme_results, 1):
                    theme = theme_result.get('theme', {})
                    stats = theme_result.get('stats', {})

                    summary_data.append({
                        '序号': idx,
                        '主题名称': theme.get('theme_name', ''),
                        '副标题': theme.get('slogan', ''),
                        '情境描述': theme.get('description', ''),
                        '预期氛围': theme.get('target_vibe', ''),
                        '向量召回': stats.get('vector_count', 0),
                        'BM25召回': stats.get('bm25_count', 0),
                        '最终推荐': stats.get('final_count', 0)
                    })

                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='主题汇总', index=False)

                # 2. 每个主题的详细推荐表
                for idx, theme_result in enumerate(theme_results, 1):
                    theme = theme_result.get('theme', {})
                    books = theme_result.get('books', [])
                    theme_name = theme.get('theme_name', f'主题{idx}')

                    # 限制sheet名称长度（Excel限制31字符）
                    sheet_name = f"{idx}.{theme_name[:20]}"
                    # 清理sheet名称中的非法字符
                    sheet_name = re.sub(r'[\\/*?:[\]]', '', sheet_name)

                    # 构建该主题的推荐数据
                    data = []
                    for i, book in enumerate(books, 1):
                        tags = self._parse_tags(book.get('tags_json', ''))

                        row_data = {
                            '序号': i,
                            '书名': book.get('title', ''),
                            '作者': book.get('author', ''),
                            '索书号': book.get('call_no', ''),
                            'RRF得分': round(book.get('rrf_score', 0), 4),
                            '向量得分': round(book.get('vector_score', 0), 4),
                            'BM25得分': round(book.get('bm25_score', 0), 4),
                            '来源': '+'.join(book.get('sources', [])),
                            '向量排名': book.get('vector_rank', ''),
                            'BM25排名': book.get('bm25_rank', ''),
                        }

                        # 重排序得分（如果有）
                        if 'rerank_score' in book:
                            row_data['重排序得分'] = round(book.get('rerank_score', 0), 4)

                        # 添加books表的额外字段
                        book_id = book.get('book_id')
                        if book_id and book_id in books_details:
                            detail = books_details[book_id]
                            for field in DOUBAN_FIELDS:
                                if field == 'barcode':
                                    row_data['条码'] = detail.get(field, '')
                                else:
                                    row_data[field] = detail.get(field, '')

                        data.append(row_data)

                    df = pd.DataFrame(data)

                    # 导出到sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # 设置列宽
                    from openpyxl.utils import get_column_letter
                    worksheet = writer.sheets[sheet_name]

                    # 基础列宽度
                    base_widths = [5, 35, 15, 12, 10, 10, 10, 10, 10, 10]
                    if any('rerank_score' in b for b in books):
                        base_widths.append(10)  # 重排序得分列

                    # douban字段列宽度
                    douban_widths = {
                        'douban_url': 40,
                        'douban_title': 30,
                        'douban_subtitle': 25,
                        'douban_original_title': 30,
                        'douban_author': 20,
                        'douban_translator': 20,
                        'douban_publisher': 20,
                        'douban_producer': 20,
                        'douban_series': 25,
                        'douban_series_link': 40,
                        'douban_price': 10,
                        'douban_isbn': 15,
                        'douban_pages': 10,
                        'douban_binding': 12,
                        'douban_pub_year': 12,
                        'douban_rating': 10,
                        'douban_rating_count': 12,
                        'douban_summary': 50,
                        'douban_author_intro': 40,
                        'douban_catalog': 40,
                        'douban_cover_image': 40
                    }

                    # 按DOUBAN_FIELDS顺序构建完整列宽列表
                    all_widths = base_widths[:]
                    for field in DOUBAN_FIELDS:
                        if field == 'barcode':
                            continue
                        all_widths.append(douban_widths.get(field, 15))

                    # 应用列宽
                    for i, width in enumerate(all_widths, 1):
                        col_letter = get_column_letter(i)
                        worksheet.column_dimensions[col_letter].width = width

                # 3. 元信息表
                meta_data = [
                    ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                    ['主题数量', len(theme_results)],
                    ['总推荐书籍', sum(len(t.get('books', [])) for t in theme_results)],
                    ['检索方法', '混合检索（向量 + BM25 + RRF融合）']
                ]
                meta_df = pd.DataFrame(meta_data, columns=['项目', '内容'])
                meta_df.to_excel(writer, sheet_name='元信息', index=False)

            logger.info(f"批量导出成功: {output_file}")
            return str(output_file)

        except Exception as e:
            logger.error(f"批量导出失败: {str(e)}")
            raise
