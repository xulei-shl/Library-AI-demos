"""
Excel操作工具模块
提供Excel文件读写和格式处理功能
"""
import pandas as pd
from pathlib import Path
from typing import List, Optional, Dict, Any
import time
import os
import re

from .logger import logger
from ..config.settings import (
    MAX_RETRY_TIMES, RETRY_DELAY, EXCEL_SHEET_NAME, 
    SHEET_PRIORITY, TARGET_FILE_PATH, FILE_PATTERNS, get_target_file_name
)
from .config_mapper import ConfigMapper


class ExcelUtils:
    """Excel操作工具类"""
    
    @staticmethod
    def find_column_by_name_or_letter(df: pd.DataFrame, column_name: str, column_letter: Optional[str] = None) -> Optional[str]:
        """
        按列名查找列，如果找不到则使用字母列号作为回退
        
        Args:
            df: DataFrame对象
            column_name: 优先查找的列名
            column_letter: 回退使用的字母列号（如'A', 'B', 'I'等）
            
        Returns:
            找到的列名，如果都找不到返回None
        """
        try:
            # 1. 优先按列名查找
            if column_name in df.columns:
                logger.debug(f"按列名找到列: {column_name}")
                return column_name
            
            # 2. 如果没有提供字母列号，直接返回None
            if not column_letter:
                logger.warning(f"未找到列 '{column_name}' 且未提供字母列号回退选项")
                return None
            
            # 3. 使用字母列号作为回退
            # 将字母列号转换为数字索引（A=0, B=1, C=2, ...）
            column_index = ExcelUtils._letter_to_index(column_letter)
            
            # 检查索引是否在有效范围内
            if 0 <= column_index < len(df.columns):
                actual_column_name = df.columns[column_index]
                logger.info(f"按列名 '{column_name}' 未找到，使用字母列号 '{column_letter}' (索引{column_index}) 回退到列: {actual_column_name}")
                return str(actual_column_name)
            else:
                logger.error(f"字母列号 '{column_letter}' 对应的索引 {column_index} 超出范围（共 {len(df.columns)} 列）")
                logger.debug(f"可用列: {list(df.columns)}")
                return None
                
        except Exception as e:
            logger.exception(f"查找列失败: column_name={column_name}, column_letter={column_letter}, 错误: {str(e)}")
            return None
    
    @staticmethod
    def _letter_to_index(letter: str) -> int:
        """
        将Excel列字母转换为数字索引
        
        Args:
            letter: 列字母（如'A', 'B', 'AA'等）
            
        Returns:
            对应的数字索引（A=0, B=1, ...）
        """
        try:
            letter = letter.upper()
            result = 0
            for char in letter:
                result = result * 26 + (ord(char) - ord('A') + 1)
            return result - 1  # 转换为0基索引
        except Exception as e:
            logger.exception(f"字母列号转换失败: {letter}, 错误: {str(e)}")
            return -1
    
    @staticmethod
    def _index_to_letter(index: int) -> str:
        """
        将数字索引转换为Excel列字母
        
        Args:
            index: 数字索引（0=A, 1=B, 8=I, ...）
            
        Returns:
            对应的列字母（如'A', 'B', 'I'等）
        """
        try:
            if index < 0:
                raise ValueError(f"索引不能为负数: {index}")
            
            result = ''
            index += 1  # 转换为1基索引
            
            while index > 0:
                index -= 1  # 调整为0基索引进行计算
                result = chr(ord('A') + index % 26) + result
                index //= 26
            
            return result
        except Exception as e:
            logger.exception(f"索引转字母失败: {index}, 错误: {str(e)}")
            return f"列{index + 1}"
    
    @staticmethod
    def get_optimal_sheet_name(file_path: Path) -> Optional[str]:
        """
        获取最优的sheet名称
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            最优的sheet名称，如果获取失败返回None
        """
        try:
            import openpyxl
            
            # 读取Excel文件的所有sheet名称
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            logger.debug(f"Excel文件 {file_path.name} 包含sheet: {sheet_names}")
            
            # 如果只有一个sheet，直接使用
            if len(sheet_names) == 1:
                logger.info(f"Excel文件只有1个sheet，使用: {sheet_names[0]}")
                return sheet_names[0]
            
            # 按优先级查找sheet
            for priority_sheet in SHEET_PRIORITY:
                if priority_sheet in sheet_names:
                    logger.info(f"找到优先级sheet: {priority_sheet}")
                    return priority_sheet
            
            # 如果都没找到，使用第一个sheet
            logger.warning(f"未找到优先级sheet，使用第一个sheet: {sheet_names[0]}")
            return sheet_names[0]
            
        except Exception as e:
            logger.exception(f"获取sheet名称失败: {file_path}, 错误: {str(e)}")
            return None
    
    @staticmethod
    def get_file_type_from_name(filename: str) -> str:
        """
        根据文件名确定文件类型
        
        Args:
            filename: 文件名
            
        Returns:
            文件类型标识
        """
        for file_type, pattern in FILE_PATTERNS.items():
            if pattern in filename:
                return file_type
        return 'default'
    
    @staticmethod
    def apply_column_mapping(columns: List[str], file_type: str) -> List[str]:
        """
        应用列名映射（使用ConfigMapper）
        
        Args:
            columns: 原始列名列表
            file_type: 文件类型
            
        Returns:
            映射后的列名列表
        """
        try:
            # 创建一个有数据的临时DataFrame来使用ConfigMapper
            temp_data = {col: [None] for col in columns}  # 每列一个None值
            temp_df = pd.DataFrame(temp_data)
            mapped_df, _ = ConfigMapper.apply_column_mapping(temp_df, file_type)
            return list(mapped_df.columns)
        except Exception as e:
            logger.exception(f"ExcelUtils.apply_column_mapping 失败: {str(e)}")
            return columns  # 失败时返回原始列名
    
    @staticmethod
    def get_target_file_headers() -> Optional[List[str]]:
        """
        获取目标文件的表头（应用动态列名映射）
        
        Returns:
            目标文件表头列表，如果获取失败返回None
        """
        try:
            target_path = Path(TARGET_FILE_PATH)
            if not target_path.exists():
                logger.error(f"目标文件不存在: {target_path}")
                return None
            
            headers = ExcelUtils.get_excel_columns(target_path, EXCEL_SHEET_NAME)
            if headers:
                logger.debug(f"目标文件原始表头: {headers}")
                # 对目标文件表头也应用动态列名映射
                mapped_headers = ExcelUtils.apply_column_mapping(headers, 'default')
                logger.debug(f"目标文件映射后表头: {mapped_headers}")
                return mapped_headers
            return headers
            
        except Exception as e:
            logger.exception(f"获取目标文件表头失败: {str(e)}")
            return None
    
    @staticmethod
    def validate_headers_with_mapping(file_path: Path, sheet_name: str, filename: str) -> bool:
        """
        验证表头是否与目标文件匹配（包容性验证，支持列名映射）
        
        Args:
            file_path: Excel文件路径
            sheet_name: sheet名称
            filename: 文件名（用于确定文件类型）
            
        Returns:
            表头是否匹配（包容性验证）
        """
        try:
            # 获取源文件表头
            source_headers = ExcelUtils.get_excel_columns(file_path, sheet_name)
            if not source_headers:
                logger.error(f"无法获取源文件表头: {file_path}, sheet: {sheet_name}")
                return False
            
            # 获取目标文件表头
            target_headers = ExcelUtils.get_target_file_headers()
            if not target_headers:
                logger.error("无法获取目标文件表头")
                return False
            
            # 确定文件类型并应用列名映射
            file_type = ExcelUtils.get_file_type_from_name(filename)
            mapped_headers = ExcelUtils.apply_column_mapping(source_headers, file_type)
            
            # 包容性验证：检查源文件的列是否能在目标文件中找到对应位置
            source_set = set(mapped_headers)
            target_set = set(target_headers)
            
            # 计算匹配的列
            matched_columns = source_set & target_set
            missing_in_source = target_set - source_set
            extra_in_source = source_set - target_set
            
            # 如果有匹配的列，就允许处理
            if matched_columns:
                logger.info(f"表头验证通过: {filename}")
                logger.info(f"匹配列数: {len(matched_columns)}/{len(target_headers)}")
                
                if missing_in_source:
                    logger.info(f"源文件缺少列（将填充空值）: {list(missing_in_source)}")
                
                if extra_in_source:
                    logger.info(f"源文件多余列（将忽略）: {list(extra_in_source)}")
                
                if file_type != 'default':
                    logger.debug(f"应用了 {file_type} 类型的列名映射")
                
                return True
            else:
                # 完全没有匹配的列，拒绝处理
                logger.warning(f"表头验证失败: {filename}")
                logger.warning("源文件与目标文件没有任何匹配的列")
                logger.warning(f"源文件列: {list(source_set)}")
                logger.warning(f"目标文件列: {list(target_set)}")
                logger.info(f"跳过处理文件: {filename}")
                return False
                
        except Exception as e:
            logger.exception(f"表头验证失败: {file_path}, 错误: {str(e)}")
            return False
    
    @staticmethod
    def read_excel_file(file_path: Path, sheet_name: Optional[str] = None, filename: Optional[str] = None) -> Optional[pd.DataFrame]:
        """
        读取Excel文件（支持智能sheet选择和表头验证）
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称（可选，如果不提供则智能选择）
            filename: 文件名（用于表头验证，如果不提供则从路径提取）
            
        Returns:
            DataFrame或None（如果读取失败或表头验证失败）
        """
        try:
            logger.info(f"开始读取Excel文件: {file_path}")
            
            # 检查文件是否存在
            if not file_path.exists():
                logger.error(f"文件不存在: {file_path}")
                return None
            
            # 检查文件是否被占用
            if not ExcelUtils._is_file_accessible(file_path):
                logger.warning(f"文件被占用，等待释放: {file_path}")
                time.sleep(RETRY_DELAY)
                if not ExcelUtils._is_file_accessible(file_path):
                    logger.error(f"文件持续被占用，无法读取: {file_path}")
                    return None
            
            # 如果没有提供filename，从路径中提取
            if filename is None:
                filename = file_path.name
            
            # 智能选择sheet名称
            if sheet_name is None:
                sheet_name = ExcelUtils.get_optimal_sheet_name(file_path)
                if sheet_name is None:
                    logger.error(f"无法确定要读取的sheet: {file_path}")
                    return None
            
            # 验证表头
            if not ExcelUtils.validate_headers_with_mapping(file_path, sheet_name, filename):
                logger.warning(f"表头验证失败，跳过文件: {filename}")
                return None
            
            # 读取Excel文件
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            logger.info(f"成功读取文件，数据行数: {len(df)}, 列数: {len(df.columns)}")
            logger.debug(f"使用sheet: {sheet_name}, 列名: {list(df.columns)}")
            
            return df
            
        except Exception as e:
            logger.exception(f"读取Excel文件失败: {file_path}, 错误: {str(e)}")
            return None
    
    @staticmethod
    def write_excel_file(df: pd.DataFrame, file_path: Path, sheet_name: str = EXCEL_SHEET_NAME) -> bool:
        """
        写入Excel文件
        
        Args:
            df: 要写入的DataFrame
            file_path: 目标文件路径
            sheet_name: 工作表名称
            
        Returns:
            是否写入成功
        """
        try:
            logger.info(f"开始写入Excel文件: {file_path}")
            
            # 确保目录存在
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 写入文件
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"成功写入文件，数据行数: {len(df)}")
            return True
            
        except Exception as e:
            logger.exception(f"写入Excel文件失败: {file_path}, 错误: {str(e)}")
            return False
    
    @staticmethod
    def append_to_excel_file(df: pd.DataFrame, file_path: Path, sheet_name: str = EXCEL_SHEET_NAME) -> bool:
        """
        追加数据到Excel文件（保留原有公式）
        
        Args:
            df: 要追加的DataFrame
            file_path: 目标文件路径
            sheet_name: 工作表名称
            
        Returns:
            是否追加成功
        """
        try:
            import openpyxl
            
            logger.info(f"开始追加数据到Excel文件: {file_path}")

            # 检查文件是否存在
            if not file_path.exists():
                logger.error(f"目标文件不存在: {file_path}")
                return False

            # 加载现有工作簿
            workbook = openpyxl.load_workbook(file_path)
            
            # 获取目标工作表
            if sheet_name not in workbook.sheetnames:
                logger.error(f"工作表 '{sheet_name}' 不存在于文件 {file_path}")
                workbook.close()
                return False
            
            worksheet = workbook[sheet_name]
            
            # 找到现有数据的最后一行
            last_row = worksheet.max_row
            logger.debug(f"当前工作表最后一行: {last_row}")
            
            # 将DataFrame数据追加到工作表
            start_row = last_row + 1
            for row_idx, (_, row_data) in enumerate(df.iterrows()):
                current_row = start_row + row_idx
                for col_idx, value in enumerate(row_data, start=1):
                    # 处理NaN值
                    if pd.isna(value):
                        cell_value = None
                    else:
                        cell_value = value
                    
                    worksheet.cell(row=current_row, column=col_idx, value=cell_value)
            
            logger.info(f"追加了 {len(df)} 行数据，从第 {start_row} 行开始")
            
            # 保存工作簿
            workbook.save(file_path)
            workbook.close()
            
            logger.info("数据追加成功，原有公式已保留")
            return True
            
        except Exception as e:
            logger.exception(f"追加数据到Excel文件失败: {file_path}, 错误: {str(e)}")
            return False
    
    @staticmethod
    def get_excel_columns(file_path: Path, sheet_name: str = EXCEL_SHEET_NAME) -> Optional[List[str]]:
        """
        获取Excel文件的列名
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            列名列表或None
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
            return list(df.columns)
        except Exception as e:
            logger.exception(f"获取Excel列名失败: {file_path}, 错误: {str(e)}")
            return None
    
    @staticmethod
    def _is_file_accessible(file_path: Path) -> bool:
        """
        检查文件是否可访问（未被占用）
        
        Args:
            file_path: 文件路径
            
        Returns:
            是否可访问
        """
        try:
            with open(file_path, 'r+b'):
                pass
            return True
        except (IOError, OSError):
            return False
    
    @staticmethod
    def find_excel_files(directory: Path) -> List[Path]:
        """
        查找目录下的所有Excel文件
        
        Args:
            directory: 搜索目录
            
        Returns:
            Excel文件路径列表
        """
        excel_files = []
        try:
            for pattern in ['*.xlsx', '*.xls']:
                excel_files.extend(directory.glob(pattern))
            
            logger.info(f"在目录 {directory} 中找到 {len(excel_files)} 个Excel文件")
            for file in excel_files:
                logger.debug(f"找到Excel文件: {file.name}")
                
        except Exception as e:
            logger.exception(f"搜索Excel文件失败: {directory}, 错误: {str(e)}")
        
        return excel_files