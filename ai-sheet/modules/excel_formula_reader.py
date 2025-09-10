"""
Excel文件读取器
支持.xlsx和.xls格式，提供分批读取、数据类型识别和结果写入功能
"""

import logging
import pandas as pd
import numpy as np
from typing import Any, Dict, List, Optional, Tuple, Union
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import xlrd
import time

# 配置日志
logger = logging.getLogger(__name__)


class ExcelFormulaReader:
    """Excel文件读取器"""
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls']
        self.current_file = None
        self.current_workbook = None
        self.current_sheet = None
        
        logger.info("Excel文件读取器初始化完成")
    
    def get_file_info(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """获取Excel文件基本信息"""
        try:
            file_path = Path(file_path)
            
            if not file_path.exists():
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            if file_path.suffix.lower() not in self.supported_formats:
                raise ValueError(f"不支持的文件格式: {file_path.suffix}")
            
            info = {
                'file_path': str(file_path),
                'file_size': file_path.stat().st_size,
                'file_format': file_path.suffix.lower(),
                'sheets': [],
                'total_rows': 0,
                'total_columns': 0,
                'current_sheet': sheet_name
            }
            
            if file_path.suffix.lower() == '.xlsx':
                # 处理.xlsx文件
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                info['sheets'] = workbook.sheetnames
                
                # 获取指定工作表或第一个工作表
                target_sheet = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
                worksheet = workbook[target_sheet]
                
                info['current_sheet'] = target_sheet
                info['total_rows'] = worksheet.max_row
                info['total_columns'] = worksheet.max_column
                
                workbook.close()
                
            elif file_path.suffix.lower() == '.xls':
                # 处理.xls文件
                workbook = xlrd.open_workbook(file_path)
                info['sheets'] = workbook.sheet_names()
                
                # 获取指定工作表或第一个工作表
                target_sheet = sheet_name if sheet_name in workbook.sheet_names() else workbook.sheet_names()[0]
                worksheet = workbook.sheet_by_name(target_sheet)
                
                info['current_sheet'] = target_sheet
                info['total_rows'] = worksheet.nrows
                info['total_columns'] = worksheet.ncols
            
            logger.info(f"文件信息获取成功: {info['total_rows']}行 x {info['total_columns']}列")
            return info
            
        except Exception as e:
            error_msg = f"获取文件信息失败: {str(e)}"
            logger.error(error_msg)
            raise Exception(error_msg)
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """获取工作表名称列表"""
        try:
            file_path = Path(file_path)
            
            if file_path.suffix.lower() == '.xlsx':
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                sheets = workbook.sheetnames
                workbook.close()
                return sheets
            elif file_path.suffix.lower() == '.xls':
                workbook = xlrd.open_workbook(file_path)
                return workbook.sheet_names()
            else:
                raise ValueError(f"不支持的文件格式: {file_path.suffix}")
                
        except Exception as e:
            logger.error(f"获取工作表名称失败: {str(e)}")
            return []
    
    def read_sample_data(self, file_path: str, max_rows: int = 10, 
                        sheet_name: Optional[str] = None) -> Optional[pd.DataFrame]:
        """读取样本数据（前N行）"""
        try:
            logger.info(f"读取样本数据: {file_path}, 最大行数: {max_rows}")
            
            # 使用pandas读取，更简单高效
            if Path(file_path).suffix.lower() == '.xlsx':
                data = pd.read_excel(
                    file_path, 
                    sheet_name=sheet_name,
                    nrows=max_rows,
                    engine='openpyxl'
                )
            else:
                data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    nrows=max_rows,
                    engine='xlrd'
                )
            
            # 确保返回DataFrame而不是dict
            if isinstance(data, dict):
                # 如果返回的是dict（多个sheet），取第一个
                data = list(data.values())[0]
            
            # 数据预处理
            data = self._preprocess_data(data)
            
            logger.info(f"样本数据读取成功: {len(data)}行 x {len(data.columns)}列")
            return data
            
        except Exception as e:
            error_msg = f"读取样本数据失败: {str(e)}"
            logger.error(error_msg)
            return None
    
    def read_batch_data(self, file_path: str, start_row: int, end_row: int,
                       sheet_name: Optional[str] = None) -> Optional[pd.DataFrame]:
        """分批读取数据"""
        try:
            logger.debug(f"读取批次数据: 行 {start_row+1}-{end_row}")
            
            # 计算实际要读取的行数
            skip_rows = start_row
            nrows = end_row - start_row
            
            if Path(file_path).suffix.lower() == '.xlsx':
                data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    skiprows=skip_rows,
                    nrows=nrows,
                    engine='openpyxl'
                )
            else:
                data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    skiprows=skip_rows,
                    nrows=nrows,
                    engine='xlrd'
                )
            
            # 确保返回DataFrame而不是dict
            if isinstance(data, dict):
                # 如果返回的是dict（多个sheet），取第一个
                data = list(data.values())[0]
            
            # 数据预处理
            data = self._preprocess_data(data)
            
            logger.debug(f"批次数据读取成功: {len(data)}行")
            return data
            
        except Exception as e:
            logger.error(f"读取批次数据失败: {str(e)}")
            return None
    
    def read_full_data(self, file_path: str, sheet_name: Optional[str] = None) -> Optional[pd.DataFrame]:
        """读取完整数据"""
        try:
            logger.info(f"读取完整数据: {file_path}")
            
            if Path(file_path).suffix.lower() == '.xlsx':
                data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    engine='openpyxl'
                )
            else:
                data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    engine='xlrd'
                )
            
            # 确保返回DataFrame而不是dict
            if isinstance(data, dict):
                # 如果返回的是dict（多个sheet），取第一个
                data = list(data.values())[0]
            
            # 数据预处理
            data = self._preprocess_data(data)
            
            logger.info(f"完整数据读取成功: {len(data)}行 x {len(data.columns)}列")
            return data
            
        except Exception as e:
            error_msg = f"读取完整数据失败: {str(e)}"
            logger.error(error_msg)
            return None
    
    def _preprocess_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """数据预处理"""
        try:
            # 处理列名
            data.columns = [str(col).strip() for col in data.columns]
            
            # 处理重复列名
            cols = pd.Series(data.columns)
            for dup in cols[cols.duplicated()].unique():
                cols[cols[cols == dup].index.values.tolist()] = [dup + '_' + str(i) if i != 0 else dup 
                                                                for i in range(sum(cols == dup))]
            data.columns = cols
            
            # 处理数据类型
            for col in data.columns:
                # 尝试转换数值类型
                if data[col].dtype == 'object':
                    # 尝试转换为数值
                    numeric_data = pd.to_numeric(data[col], errors='coerce')
                    if not numeric_data.isna().all():
                        # 如果转换后不全是NaN，则使用数值类型
                        data[col] = numeric_data
            
            # 重置索引
            data.reset_index(drop=True, inplace=True)
            
            return data
            
        except Exception as e:
            logger.warning(f"数据预处理失败: {str(e)}")
            return data
    
    def save_data(self, file_path: str, data: pd.DataFrame, 
                 sheet_name: Optional[str] = None, 
                 overwrite_strategy: str = "overwrite") -> bool:
        """保存数据到Excel文件"""
        try:
            logger.info(f"保存数据到文件: {file_path}")
            file_path = Path(file_path)
            
            if overwrite_strategy == "overwrite":
                # 直接覆盖整个文件
                if file_path.suffix.lower() == '.xlsx':
                    self._save_with_formula_support(file_path, data, sheet_name or 'Sheet1')
                else:
                    # 对于.xls文件，转换为.xlsx
                    new_path = file_path.with_suffix('.xlsx')
                    self._save_with_formula_support(new_path, data, sheet_name or 'Sheet1')
                    logger.info(f"文件已转换为.xlsx格式: {new_path}")
                    
            elif overwrite_strategy == "append":
                # 新增列到现有工作表
                self._add_new_column_with_formula_support(file_path, data, sheet_name)
                
            elif overwrite_strategy == "new_sheet":
                # 创建新工作表
                self._add_new_sheet_with_formula_support(file_path, data, sheet_name)
            
            logger.info("数据保存成功")
            return True
            
        except Exception as e:
            error_msg = f"保存数据失败: {str(e)}"
            logger.error(error_msg)
            return False
    
    def _add_new_column(self, file_path: Path, data: pd.DataFrame, sheet_name: Optional[str]):
        """在现有工作表中新增列（已弃用，使用_add_new_column_with_formula_support代替）"""
        logger.warning("_add_new_column方法已弃用，自动转为使用_add_new_column_with_formula_support")
        return self._add_new_column_with_formula_support(file_path, data, sheet_name)
    
    def _get_unique_column_name(self, existing_columns, base_name: str) -> str:
        """生成唯一的列名，如果冲突则添加数字后缀"""
        if base_name not in existing_columns:
            return base_name
        
        counter = 1
        while f"{base_name}-{counter}" in existing_columns:
            counter += 1
        
        return f"{base_name}-{counter}"
    
    def _add_new_sheet(self, file_path: Path, data: pd.DataFrame, sheet_name: Optional[str]):
        """添加新工作表"""
        try:
            sheet_name = sheet_name or f'Sheet_{int(time.time())}'
            
            if file_path.exists() and file_path.suffix.lower() == '.xlsx':
                # 使用openpyxl添加新工作表
                workbook = openpyxl.load_workbook(file_path)
                
                # 确保工作表名称唯一
                if sheet_name in workbook.sheetnames:
                    counter = 1
                    while f"{sheet_name}_{counter}" in workbook.sheetnames:
                        counter += 1
                    sheet_name = f"{sheet_name}_{counter}"
                
                # 创建新工作表
                worksheet = workbook.create_sheet(sheet_name)
                
                # 写入数据
                for r in dataframe_to_rows(data, index=False, header=True):
                    worksheet.append(r)
                
                # 保存工作簿
                workbook.save(file_path)
                workbook.close()
                
            else:
                # 文件不存在，创建新文件
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
                    
        except Exception as e:
            logger.error(f"添加新工作表失败: {str(e)}")
            raise
    
    def update_column(self, file_path: str, column_name: str, column_data: pd.Series,
                     sheet_name: Optional[str] = None) -> bool:
        """更新指定列的数据"""
        try:
            logger.info(f"更新列数据: {column_name}")
            
            # 读取现有数据
            existing_data = self.read_full_data(file_path, sheet_name)
            if existing_data is None:
                return False
            
            # 更新列数据
            if len(column_data) <= len(existing_data):
                existing_data.loc[:len(column_data)-1, column_name] = column_data.values
            else:
                # 如果新数据更长，需要扩展DataFrame
                existing_data = existing_data.reindex(range(len(column_data)))
                existing_data[column_name] = column_data.values
            
            # 保存更新后的数据
            return self.save_data(file_path, existing_data, sheet_name, "overwrite")
            
        except Exception as e:
            logger.error(f"更新列数据失败: {str(e)}")
            return False
    
    def get_column_names(self, file_path: str, sheet_name: Optional[str] = None) -> List[str]:
        """获取列名列表"""
        try:
            # 只读取第一行来获取列名
            if Path(file_path).suffix.lower() == '.xlsx':
                data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    nrows=1,
                    engine='openpyxl'
                )
            else:
                data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    nrows=1,
                    engine='xlrd'
                )
            
            # 确保返回DataFrame而不是dict
            if isinstance(data, dict):
                # 如果返回的是dict（多个sheet），取第一个
                data = list(data.values())[0]
            
            return data.columns.tolist()
            
        except Exception as e:
            logger.error(f"获取列名失败: {str(e)}")
            return []
    
    def validate_file(self, file_path: str) -> Tuple[bool, str]:
        """验证Excel文件"""
        try:
            file_path = Path(file_path)
            
            # 检查文件是否存在
            if not file_path.exists():
                return False, "文件不存在"
            
            # 检查文件格式
            if file_path.suffix.lower() not in self.supported_formats:
                return False, f"不支持的文件格式: {file_path.suffix}"
            
            # 检查文件是否可读
            try:
                info = self.get_file_info(str(file_path))
                if info['total_rows'] == 0:
                    return False, "文件为空"
                    
                return True, "文件验证通过"
                
            except Exception as e:
                return False, f"文件读取失败: {str(e)}"
                
        except Exception as e:
            return False, f"文件验证异常: {str(e)}"
    
    def close(self):
        """关闭当前打开的文件"""
        try:
            if self.current_workbook:
                if hasattr(self.current_workbook, 'close'):
                    self.current_workbook.close()
                self.current_workbook = None
                self.current_sheet = None
                self.current_file = None
                logger.info("文件已关闭")
        except Exception as e:
            logger.warning(f"关闭文件时出错: {str(e)}")
    
    def _save_with_formula_support(self, file_path: Path, data: pd.DataFrame, sheet_name: str):
        """③保存数据并支持公式单元格标记"""
        try:
            # 使用openpyxl直接操作，以便设置公式单元格类型
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            # 写入表头
            for col_idx, column_name in enumerate(data.columns, 1):
                worksheet.cell(row=1, column=col_idx, value=column_name)
            
            # 写入数据并标记公式单元格
            for row_idx, (_, row_data) in enumerate(data.iterrows(), 2):
                for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    # 检查是否为Excel公式
                    if isinstance(value, str) and value.startswith('='):
                        # 设置为公式类型
                        cell.value = value
                        cell.data_type = 'f'  # 标记为公式类型
                        logger.debug(f"设置公式单元格 {get_column_letter(col_idx)}{row_idx}: {value}")
                    else:
                        cell.value = value
            
            # 保存工作簿
            workbook.save(file_path)
            workbook.close()
            
            # 检查是否有公式单元格被标记
            formula_count = sum(1 for col in data.columns 
                              for value in data[col] 
                              if isinstance(value, str) and value.startswith('='))
            
            if formula_count > 0:
                logger.info(f"已标记 {formula_count} 个公式单元格，Excel打开后将自动重算")
                
        except Exception as e:
            logger.error(f"保存公式支持文件失败: {str(e)}")
            raise
    
    def _add_new_column_with_formula_support(self, file_path: Path, data: pd.DataFrame, sheet_name: Optional[str]):
        """在现有工作表中新增列并支持公式，保护现有公式不被清空"""
        try:
            if file_path.exists() and file_path.suffix.lower() == '.xlsx':
                # 使用 data_only=False 来保留公式
                workbook = openpyxl.load_workbook(file_path, data_only=False)
                
                # 获取目标工作表
                if sheet_name and sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                else:
                    worksheet = workbook.active
                    sheet_name = worksheet.title
                
                logger.info(f"操作工作表: {sheet_name}")
                
                # 检查并记录现有公式（用于验证保护）
                existing_formulas = self._check_existing_formulas(worksheet)
                if existing_formulas:
                    logger.info(f"检测到现有公式 {len(existing_formulas)} 个，将保护不被清空")
                
                # 获取现有数据的基本信息
                existing_rows = worksheet.max_row
                existing_cols = worksheet.max_column
                logger.info(f"现有数据: {existing_rows} 行, {existing_cols} 列")
                
                # 获取现有列名
                existing_column_names = []
                for col in range(1, existing_cols + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    existing_column_names.append(str(cell_value) if cell_value is not None else f"Column{col}")
                
                logger.info(f"现有列名: {existing_column_names}")
                
                # 确定要添加的新列
                columns_to_add = list(data.columns)
                logger.info(f"待处理的列: {columns_to_add}")
                
                # 处理每个新列
                next_col = existing_cols + 1
                actually_added_columns = []
                
                for new_col in columns_to_add:
                    # 生成唯一列名
                    final_col_name = self._get_unique_column_name(existing_column_names, new_col)
                    
                    # 检查是否真的需要新增
                    if final_col_name in existing_column_names:
                        logger.info(f"列 '{new_col}' 已存在，跳过新增")
                        continue
                    
                    logger.info(f"新增列: {new_col} -> {final_col_name}")
                    
                    # 写入新列标题
                    worksheet.cell(row=1, column=next_col, value=final_col_name)
                    
                    # 写入新列数据
                    data_rows_written = 0
                    for row_idx, value in enumerate(data[new_col], 2):  # 从第2行开始
                        cell = worksheet.cell(row=row_idx, column=next_col)
                        
                        # 检查是否为Excel公式
                        if isinstance(value, str) and value.startswith('='):
                            cell.value = value
                            cell.data_type = 'f'  # 标记为公式类型
                            logger.debug(f"设置公式单元格 {openpyxl.utils.get_column_letter(next_col)}{row_idx}: {value}")
                        else:
                            cell.value = value
                        
                        data_rows_written += 1
                    
                    # 更新状态
                    existing_column_names.append(final_col_name)
                    actually_added_columns.append(final_col_name)
                    logger.info(f"成功新增列: {final_col_name} 在第 {next_col} 列，数据行数: {data_rows_written}")
                    next_col += 1
                
                if not actually_added_columns:
                    logger.info("没有新增任何列")
                else:
                    logger.info(f"总共新增了 {len(actually_added_columns)} 列: {actually_added_columns}")
                
                # 保存工作簿（关键：这里会保留所有现有的公式）
                workbook.save(file_path)
                
                # 验证现有公式是否被保护
                if existing_formulas:
                    logger.info("验证现有公式保护情况...")
                    # 重新加载验证
                    verify_workbook = openpyxl.load_workbook(file_path, data_only=False)
                    verify_worksheet = verify_workbook[sheet_name]
                    preserved_formulas = self._check_existing_formulas(verify_worksheet)
                    
                    if len(preserved_formulas) >= len(existing_formulas):
                        logger.info(f"✓ 现有公式保护成功，保留了 {len(preserved_formulas)} 个公式")
                    else:
                        logger.warning(f"⚠ 可能有公式丢失，原有 {len(existing_formulas)} 个，现有 {len(preserved_formulas)} 个")
                    
                    verify_workbook.close()
                
                workbook.close()
                
                # 统计新添加的公式数量
                formula_count = sum(1 for col in data.columns 
                                for value in data[col] 
                                if isinstance(value, str) and value.startswith('='))
                
                if formula_count > 0:
                    logger.info(f"已添加 {formula_count} 个新公式单元格")
                    
            else:
                # 文件不存在，直接创建
                logger.info(f"文件不存在，创建新文件: {file_path}")
                self._save_with_formula_support(file_path, data, sheet_name or 'Sheet1')
                    
        except Exception as e:
            logger.error(f"新增列失败: {str(e)}")
            raise

    def _check_existing_formulas(self, worksheet) -> List[Dict[str, Any]]:
        """检查工作表中现有的公式单元格"""
        try:
            formulas = []
            
            # 遍历所有单元格，查找公式
            for row in worksheet.iter_rows():
                for cell in row:
                    # 检查单元格是否包含公式
                    if cell.data_type == 'f' and cell.value and str(cell.value).startswith('='):
                        formulas.append({
                            'address': f"{openpyxl.utils.get_column_letter(cell.column)}{cell.row}",
                            'formula': cell.value,
                            'row': cell.row,
                            'column': cell.column
                        })
            
            if formulas:
                logger.info(f"检测到 {len(formulas)} 个现有公式单元格")
                # 记录前几个公式作为示例
                for i, formula_info in enumerate(formulas[:3]):
                    logger.debug(f"现有公式 {i+1}: {formula_info['address']} = {formula_info['formula']}")
                if len(formulas) > 3:
                    logger.debug(f"... 还有 {len(formulas) - 3} 个现有公式")
            
            return formulas
            
        except Exception as e:
            logger.warning(f"检查现有公式失败: {str(e)}")
            return []

    def _get_unique_column_name(self, existing_columns: List[str], base_name: str) -> str:
        """生成唯一的列名，如果冲突则添加数字后缀"""
        if base_name not in existing_columns:
            return base_name
        
        counter = 1
        while f"{base_name}-{counter}" in existing_columns:
            counter += 1
        
        unique_name = f"{base_name}-{counter}"
        logger.debug(f"列名冲突解决: {base_name} -> {unique_name}")
        return unique_name
    
    def _add_new_sheet_with_formula_support(self, file_path: Path, data: pd.DataFrame, sheet_name: Optional[str]):
        """添加新工作表并支持公式"""
        try:
            sheet_name = sheet_name or f'Sheet_{int(time.time())}'
            
            if file_path.exists() and file_path.suffix.lower() == '.xlsx':
                # 使用openpyxl添加新工作表
                workbook = openpyxl.load_workbook(file_path)
                
                # 确保工作表名称唯一
                if sheet_name in workbook.sheetnames:
                    counter = 1
                    while f"{sheet_name}_{counter}" in workbook.sheetnames:
                        counter += 1
                    sheet_name = f"{sheet_name}_{counter}"
                
                # 创建新工作表
                worksheet = workbook.create_sheet(sheet_name)
                
                # 写入表头
                for col_idx, column_name in enumerate(data.columns, 1):
                    worksheet.cell(row=1, column=col_idx, value=column_name)
                
                # 写入数据并标记公式单元格
                for row_idx, (_, row_data) in enumerate(data.iterrows(), 2):
                    for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        
                        # 检查是否为Excel公式
                        if isinstance(value, str) and value.startswith('='):
                            cell.value = value
                            cell.data_type = 'f'  # 标记为公式类型
                        else:
                            cell.value = value
                
                # 保存工作簿
                workbook.save(file_path)
                workbook.close()
                
            else:
                # 文件不存在，创建新文件
                self._save_with_formula_support(file_path, data, sheet_name)
                    
        except Exception as e:
            logger.error(f"添加新工作表失败: {str(e)}")
            raise

    def _check_existing_formulas(self, worksheet) -> List[Dict[str, Any]]:
        """检查工作表中现有的公式单元格"""
        try:
            formulas = []
            
            # 遍历所有单元格，查找公式
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        # 找到公式单元格
                        formulas.append({
                            'address': f"{get_column_letter(cell.column)}{cell.row}",
                            'formula': cell.value,
                            'row': cell.row,
                            'column': cell.column
                        })
            
            if formulas:
                logger.info(f"检测到 {len(formulas)} 个现有公式单元格")
                # 记录前几个公式作为示例
                for i, formula_info in enumerate(formulas[:3]):
                    logger.debug(f"公式 {i+1}: {formula_info['address']} = {formula_info['formula']}")
                if len(formulas) > 3:
                    logger.debug(f"... 还有 {len(formulas) - 3} 个公式")
            
            return formulas
            
        except Exception as e:
            logger.warning(f"检查现有公式失败: {str(e)}")
            return []

    def __del__(self):
        """析构函数"""
        self.close()