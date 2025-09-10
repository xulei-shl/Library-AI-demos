"""
Excel处理模块 - 负责Excel文件的读取、写入和数据管理
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
from typing import List, Dict, Any, Optional, Tuple
import logging


class ExcelProcessor:
    """Excel文件处理器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.workbook = None
        self.worksheet = None
        self.file_path = None
        
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """获取Excel文件信息"""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
                
            # 使用pandas快速获取基本信息
            df = pd.read_excel(file_path, nrows=0)  # 只读取表头
            
            # 使用openpyxl获取详细信息
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            info = {
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames,
                'row_count': ws.max_row,
                'column_count': ws.max_column,
                'columns': list(df.columns) if not df.empty else []
            }
            
            wb.close()
            return info
            
        except Exception as e:
            self.logger.error(f"获取文件信息失败: {e}")
            raise
            
    def load_excel(self, file_path: str) -> bool:
        """加载Excel文件"""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
                
            self.workbook = load_workbook(file_path)
            self.worksheet = self.workbook.active
            self.file_path = file_path
            
            self.logger.info(f"成功加载Excel文件: {file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"加载Excel文件失败: {e}")
            raise
            
    def get_column_data(self, columns: List[str]) -> List[Dict[str, Any]]:
        """获取指定列的数据"""
        try:
            if not self.worksheet:
                raise ValueError("未加载Excel文件")
                
            # 转换列名为列索引
            column_indices = []
            for col in columns:
                if col.isalpha():
                    # 字母列名 (A, B, C...)
                    column_indices.append(column_index_from_string(col))
                else:
                    # 数字列名
                    column_indices.append(int(col))
                    
            data_rows = []
            max_row = self.worksheet.max_row
            
            # 从第2行开始读取数据（跳过表头）
            for row_num in range(2, max_row + 1):
                row_data = {}
                combined_data = []
                
                for i, col_idx in enumerate(column_indices):
                    cell_value = self.worksheet.cell(row=row_num, column=col_idx).value
                    cell_value = str(cell_value) if cell_value is not None else ""
                    
                    row_data[columns[i]] = cell_value
                    combined_data.append(cell_value)
                
                # 使用特殊分隔符拼接多列数据
                row_data['combined'] = "|||".join(combined_data)
                row_data['row_number'] = row_num
                
                data_rows.append(row_data)
                
            self.logger.info(f"成功读取 {len(data_rows)} 行数据")
            return data_rows
            
        except Exception as e:
            self.logger.error(f"读取列数据失败: {e}")
            raise
            
    def find_result_column(self, column_name: str = "AI处理结果") -> int:
        """查找或创建结果列 - 支持自定义列名"""
        try:
            if not self.worksheet:
                raise ValueError("未加载Excel文件")
                
            # 处理列名长度限制（Excel列名最大255字符）
            if len(column_name) > 255:
                column_name = column_name[:252] + "..."
                self.logger.warning(f"列名过长，已截断为: {column_name}")
                
            # 查找指定名称的列
            for col in range(1, self.worksheet.max_column + 1):
                header_cell = self.worksheet.cell(row=1, column=col)
                if header_cell.value == column_name:
                    return col
                    
            # 如果没找到，创建新列
            result_col = self.worksheet.max_column + 1
            self.worksheet.cell(row=1, column=result_col, value=column_name)
            
            self.logger.info(f"创建结果列: {get_column_letter(result_col)} (列名: {column_name})")
            return result_col
            
        except Exception as e:
            self.logger.error(f"查找结果列失败: {e}")
            raise
            
    def write_result(self, row_number: int, result: str, result_column: int = None, column_name: str = "AI处理结果") -> bool:
        """写入处理结果 - 支持自定义列名"""
        try:
            if not self.worksheet:
                raise ValueError("未加载Excel文件")
                
            if result_column is None:
                result_column = self.find_result_column(column_name)
                
            # 写入结果
            self.worksheet.cell(row=row_number, column=result_column, value=result)
            
            return True
            
        except Exception as e:
            self.logger.error(f"写入结果失败: {e}")
            return False
            
    def check_processed(self, row_number: int, result_column: int = None, column_name: str = "AI处理结果") -> bool:
        """检查指定行是否已处理（断点续传支持） - 支持自定义列名"""
        try:
            if not self.worksheet:
                return False
                
            if result_column is None:
                result_column = self.find_result_column(column_name)
                
            cell_value = self.worksheet.cell(row=row_number, column=result_column).value
            return cell_value is not None and str(cell_value).strip() != ""
            
        except Exception as e:
            self.logger.error(f"检查处理状态失败: {e}")
            return False
            
    def save_excel(self, save_path: str = None) -> bool:
        """保存Excel文件"""
        try:
            if not self.workbook:
                raise ValueError("未加载Excel文件")
                
            save_path = save_path or self.file_path
            self.workbook.save(save_path)
            
            self.logger.info(f"成功保存Excel文件: {save_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"保存Excel文件失败: {e}")
            return False
            
    def get_processing_status(self, columns: List[str], column_name: str = "AI处理结果") -> Dict[str, int]:
        """获取处理状态统计 - 支持自定义列名"""
        try:
            if not self.worksheet:
                return {'total': 0, 'processed': 0, 'remaining': 0}
                
            result_column = self.find_result_column(column_name)
            max_row = self.worksheet.max_row
            
            total = max_row - 1  # 减去表头行
            processed = 0
            
            # 统计已处理的行数
            for row_num in range(2, max_row + 1):
                if self.check_processed(row_num, result_column, column_name):
                    processed += 1
                    
            return {
                'total': total,
                'processed': processed,
                'remaining': total - processed
            }
            
        except Exception as e:
            self.logger.error(f"获取处理状态失败: {e}")
            return {'total': 0, 'processed': 0, 'remaining': 0}
            
    def close(self):
        """关闭Excel文件"""
        try:
            if self.workbook:
                self.workbook.close()
                self.workbook = None
                self.worksheet = None
                self.file_path = None
                
        except Exception as e:
            self.logger.error(f"关闭Excel文件失败: {e}")
            
    def __enter__(self):
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


class ExcelBatchProcessor:
    """Excel批量处理器 - 支持大文件分批处理"""
    
    def __init__(self, batch_size: int = 100):
        self.batch_size = batch_size
        self.logger = logging.getLogger(__name__)
        
    def process_in_batches(self, file_path: str, columns: List[str], 
                          process_func, column_name: str = "AI处理结果", **kwargs) -> Dict[str, Any]:
        """分批处理Excel数据 - 支持自定义列名"""
        try:
            with ExcelProcessor() as processor:
                processor.load_excel(file_path)
                
                # 获取所有数据
                all_data = processor.get_column_data(columns)
                result_column = processor.find_result_column(column_name)
                
                # 分批处理
                total_rows = len(all_data)
                processed_count = 0
                success_count = 0
                failed_count = 0
                
                for i in range(0, total_rows, self.batch_size):
                    batch_data = all_data[i:i + self.batch_size]
                    
                    for row_data in batch_data:
                        row_number = row_data['row_number']
                        
                        # 检查是否已处理（断点续传） - 使用动态列名
                        if processor.check_processed(row_number, result_column, column_name):
                            processed_count += 1
                            success_count += 1
                            continue
                            
                        try:
                            # 调用处理函数
                            result = process_func(row_data, **kwargs)
                            
                            # 写入结果 - 使用动态列名
                            if processor.write_result(row_number, result, result_column, column_name):
                                success_count += 1
                            else:
                                failed_count += 1
                                
                        except Exception as e:
                            self.logger.error(f"处理第{row_number}行失败: {e}")
                            failed_count += 1
                            
                        processed_count += 1
                        
                        # 每处理3行保存一次
                        if processed_count % 3 == 0:
                            processor.save_excel()
                            
                # 最终保存
                processor.save_excel()
                
                return {
                    'total': total_rows,
                    'processed': processed_count,
                    'success': success_count,
                    'failed': failed_count
                }
                
        except Exception as e:
            self.logger.error(f"批量处理失败: {e}")
            raise